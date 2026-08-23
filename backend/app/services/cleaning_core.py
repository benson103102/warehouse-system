#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
P零售物流data 資料清洗自動化腳本
======================================
核心重點：
    依「P零售物流data_清洗文件.md」（Part 1~7）與人工分析對話紀錄整理而成的一鍵式
    清洗流程，將原始 P零售物流data.xlsx 自動分流為「乾淨資料」與「隔離資料」，並補齊
    材積（長寬高／材積cm³／貨架棧板儲位分類）與配送地址（縣市／區域）兩大類資訊，
    供下游 ABC 分組、儲位配置、區域併單分析使用。
      - 決定性：相同輸入檔重複執行，輸出逐儲存格結果完全一致（無隨機性）。
      - 已驗證：以原始檔實測，隔離列號集合、乾淨資料商品類別分派皆與人工清洗結果相符。
      - 可追溯：所有自動回填／校正都寫入「逐筆變更紀錄」；無法確定者一律標記待複核，
        不靜默混入乾淨資料。
    完整逐版更新歷程見檔案最末「更新記錄」區塊。

1. 安裝需求：不需手動安裝，直接執行即可
    腳本啟動時會自動偵測 pandas／openpyxl 是否已就緒，缺套件時會在腳本旁自動
    建立 .venv 並裝好 pandas／openpyxl／python-calamine（優先用 uv，沒有 uv
    則退回標準庫 venv），裝好後自動改用該環境重新執行，無需手動下 pip 指令。
    詳見 _bootstrap_dependencies()。若自動安裝失敗，才需要手動執行：
        uv venv
        uv pip install --python .venv\Scripts\python.exe pandas openpyxl python-calamine
    （不需要 xlsxwriter；大表輸出改用 openpyxl write_only 串流模式手動逐列寫入，
      詳見 write_output() 內部註解）

2. 用法：
    最簡單：直接雙擊同資料夾下的「執行清洗.bat」，依提示拖曳或貼上輸入檔路徑即可。

    或在終端機執行（本機請務必用 .venv 內的 python，不要打裸的 python——本機
    PATH 上的 python 對應到 Windows 市集的假殼，非互動執行下完全沒有反應，
    詳見 _bootstrap_dependencies() 上方註解）：
        .venv\Scripts\python.exe P零售物流data_清洗腳本.py <輸入檔案.xlsx> [輸出目錄]

    範例：
        .venv\Scripts\python.exe P零售物流data_清洗腳本.py "C:\path\to\P零售物流data.xlsx"
        .venv\Scripts\python.exe P零售物流data_清洗腳本.py "C:\path\to\新月份資料.xlsx" "C:\path\to\輸出資料夾"

    VS Code 內也可直接按右上角 ▶ Run 按鈕執行。

3. 輸入檔案需包含兩個工作表：
    - 訂單資料（14欄，欄位順序需與原始 P零售物流data.xlsx 一致）
    - 商品類別對應表（本腳本實際上不使用輸入檔內的這張表，
      而是採用下方寫死的 69 碼版本，因為該對照表是跨檔案的固定業務規則）

4. 輸出：<輸出目錄>/P零售物流data_清理結果.xlsx
    - 乾淨資料　　（含「出貨日期」欄＋材積清洗結果「材積_cm3／材積來源／儲位分類／
      材積待複核」＋配送地址補值欄「配送地址_補值來源」／「配送地址_待確認_flag」／
      「配送區域」——只含縣市＋區域（例：新北市內湖區），不含路名門牌，供區域併單
      策略直接依此欄分組；查無可信縣市/區域者此欄留空，同時 配送地址_待確認_flag=1；
      「配送區域」欄位置緊接在「配送地址」右邊，方便併單時直接參照）
    - 隔離資料　　（含 Excel原始列號／隔離主要原因／命中所有規則／出貨日期＋上述材積
      與配送地址補值欄位）
    - 商品類別對應表（69碼固定版本）
    - 逐筆變更紀錄（本次執行中，值被覆蓋的儲存格清單，含材積回填/校正/推估紀錄
      ＋配送地址補值紀錄）
    - 商品材積主檔（一個商品ID一列的代表長/寬/高/重量/材積(cm³)＋最長邊＋尺寸來源
      ＋儲位分類(貨架/棧板)＋分類依據，供後續 ABC 分組、貨架/棧板儲位配置等下游分析
      使用；已排除測試/範例資料商品ID）
"""

import sys
import os
import re
import subprocess
from datetime import datetime, date
from collections import defaultdict


# ============================================================
# 注意：原腳本此處有一段「依賴套件自舉」邏輯（缺 pandas/openpyxl 時自動建立 .venv
# 並安裝），供雙擊執行的獨立腳本使用。後端服務由 requirements.txt 統一管理套件，
# 故移植進後端時移除該段，直接 import pandas/numpy。
# ============================================================

import pandas as pd
import numpy as np

# Part 12：Windows 主控台在非UTF-8語系（例如繁體中文預設的 cp950/Big5）下，
# print() 遇到 log 訊息裡的符號（例如列數勾稽用的「✔」「✘」、警告用的「⚠️」）會直接
# 丟出 UnicodeEncodeError，讓整支腳本在寫log這一步當掉，而不是清洗邏輯本身出錯
# （實測：在 cp950 主控台下，程式跑到最後「列數勾稽」那行印出「✔」就崩潰，前面所有
# 清洗運算其實都已經正確跑完，只是沒能寫出最後的完成訊息）。統一把標準輸出/錯誤輸出
# 改成 UTF-8、無法編碼的字元以問號取代（不中斷執行），確保腳本在任何 Windows 語系的
# 主控台下都能跑到底，不會因為 log 訊息本身而中斷。
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")


# ============================================================
# 0. 固定設定（依 P零售物流data_清洗文件.md 各節整理，不隨輸入檔變動）
# ============================================================

ORDER_COLUMNS = [
    "系統出貨單號", "門市訂單號", "門市代碼", "配送地址",
    "商品ID", "商品編號", "商品名稱",
    "商品類別", "訂購數量", "單位", "商品-長", "商品-寬",
    "商品-高", "商品重量",
]

# 欄位屬性：是否做 NULL 字樣→真空值正規化、是否去除前後空白
# 依 清洗文件.md Part 2 §1-1：門市代碼／配送地址／商品-長/寬/高／商品重量 保留不清洗
NULL_NORMALIZE_COLUMNS = ["門市訂單號"]
TRIM_COLUMNS = ["商品編號", "商品名稱", "單位"]

# ---- 商品名稱品質（Part 17 新增）----
# 有些商品名稱是純規格代號、幾乎沒有可辨識品名文字（實測：海鮮的「(60) 10/15」「抽#2」「(82)」，
# 其中數字是規格/尺寸分級碼）。已用原始檔驗證：這 29 個商品ID在原始資料的每一列都是規格碼，
# 資料內無法補回正名，只能靠外部供應商品名對照表。判定方式：名稱去掉數字/空白/符號後，
# 剩下的「中文字或英文字母」數量少於此門檻即視為非描述性。命中者不刪除、不亂猜，改隔離待人工
# 補品名（使用者選擇）。門檻設為 2＝抓「可辨識文字≤1個」者；設為 1 則只抓完全無文字者。
NAME_MIN_MEANINGFUL_CHARS = 2
_MEANINGFUL_CHAR_RE = re.compile(r"[一-鿿A-Za-z]")

NULL_TOKENS = {"", "NULL", "N/A", "NA", "NONE", "-"}

UNIT_WHITELIST = {
    "個", "箱", "包", "PCS", "盒", "瓶", "桶", "條", "組",
    "袋", "捲", "板", "件", "支", "公斤",
}

SHIPNO_PATTERN = re.compile(r"^OU\d{8}\d{6}$")

# ---- 材積清洗（商品-長／商品-寬／商品-高／商品重量，清洗文件.md Part 9）----
DIM_COLUMNS = ["商品-長", "商品-寬", "商品-高"]
WEIGHT_COLUMN = "商品重量"
VOLUME_ALL_COLUMNS = DIM_COLUMNS + [WEIGHT_COLUMN]

# 極端值門檻：依清洗文件 Part 1 §2-2 既有分析，商品-長 > 200cm 或 商品重量 > 30kg 視為異常。
# 商品-寬／商品-高本次比照商品-長的門檻一併套用，尚未依實際資料分佈個別校準，
# 屬本次新增假設，建議之後對照實際百分位數（跑過本腳本後）重新檢視是否要拆開不同門檻。
EXTREME_MAX = {"商品-長": 200.0, "商品-寬": 200.0, "商品-高": 200.0, "商品重量": 30.0}

# 極端值疑似單位錯誤的校正候選除數（如 mm 誤填 cm → ÷10、公克誤填公斤 → ÷1000）
UNIT_CORRECTION_FACTORS = [10, 100, 1000]
# 校正候選值須落在「同商品類別有效樣本中位數」的此倍數區間內，才視為合理校正結果（可調整）
CATEGORY_PLAUSIBLE_RATIO = (1 / 3, 3)
# 商品類別至少需有此筆數的有效樣本，才採信其中位數作為校正基準（樣本太少的類別不校正，只標記）
CATEGORY_MIN_SAMPLE = 5

# ---- 長寬高缺值 → 同類別中位數推估（Part 15 新增）----
# 尺寸是「商品」屬性、同商品ID的每一筆訂單長寬高都相同，因此同ID/同名互補救不了缺值
# （實測：同商品ID眾數回填 0 筆）。唯一能自動補的是「同商品類別、其他有實測值商品」的中位數。
# 一個類別至少要有此數量的「不同商品」有實測值，才用其中位數推估缺維度的商品（避免單一樣本
# 硬套到不相關商品）。實測分佈：門檻=5 可補約863~889商品、=3約886~936、=1約993~1052。
# 使用者已選擇「最大化覆蓋」；此處取 3 為積極但不至於用單一樣本硬推的折衷，要更積極可改為 1。
# 只推估長/寬/高，不推估商品重量（本專案不使用重量）。所有推估值會標記 校正來源="類別中位數推估"，
# 並在主表「材積來源」欄標成「類別推估」，不與實測值混淆。
IMPUTE_MIN_CATEGORY_SAMPLE = 3

# ---- 材積體積門檻（Part 10 新增）----
# 長/寬/高各自的極端值判斷是單一維度分開比較，可能各自都<200cm門檻卻乘出離譜的體積。
# 依實測：本資料集正常商品材積最高約6.7萬cm³（95百分位），本次發現的異常量販包
# 最低約99.7萬cm³，門檻抓在兩者中間留安全margin，供之後依實際資料分佈調整。
VOLUME_EXTREME_MAX = 500_000.0  # cm³（約500公升／半立方公尺）

# ---- 共用可疑值偵測（Part 10 新增）----
# 只套用在「已判定為極端值」的原始值上（避免誤傷本來就常見的正常整數量測值，如10cm、0.2kg），
# 若同一個極端原始值重複出現在異常多個不同商品ID上，較可能是系統共用預設值/測試值，
# 不予單位校正、直接視為缺值。注意：這是啟發式規則，若某類商品本來就有大量SKU共用同一種
# 標準化包裝規格（例如同款式不同色的商品，物理尺寸完全一樣），也可能被誤判，門檻可再調整。
SHARED_VALUE_MIN_DISTINCT_SKU = 10  # 同一極端原始值出現在≥此數量的不同商品ID上即視為可疑

# ---- 測試/範例資料商品名稱關鍵字（Part 10 新增，供商品材積主檔排除用）----
# 範圍比 RECLASSIFY_RULES 裡的「測試商品」規則更廣，且不限於隔離資料的重分類流程，
# 直接在建立商品材積主檔時套用，涵蓋更多常見的假資料/測試資料命名慣例。
TEST_DATA_NAME_PATTERNS = [
    r"測試商品", r"測試資料", r"^測試", r"[Tt][Ee][Ss][Tt]", r"^[Dd]emo",
    r"範例商品", r"示範商品", r"^[Dd]ummy", r"^假商品", r"樣本商品",
]
_COMPILED_TEST_DATA_PATTERNS = [re.compile(p) for p in TEST_DATA_NAME_PATTERNS]

# ---- 貨架 vs 棧板 儲位分類門檻（Part 16 新增）----
# 依商品尺寸把每個商品分到「貨架區（揀貨儲位）」或「棧板區（大宗儲位）」，供下游 ABC 分組、
# 儲位配置使用。門檻為使用者依實體貨架內徑選定：材積 > 50L 或 最長邊 > 60cm → 棧板，其餘 → 貨架。
# 缺尺寸（尺寸來源=無法計算）的商品暫歸棧板待人工量測。要改門檻改這兩個常數即可。
SHELF_MAX_VOLUME_CM3 = 50_000.0  # cm³（50公升）；超過視為大宗、歸棧板
SHELF_MAX_EDGE_CM = 60.0         # cm；最長邊超過視為大件、歸棧板

# 材積清洗過程中用到的逐維度細節欄位（無法解析/負值/原始為0/極端值/共用可疑值/校正來源），
# 只在 clean_dimensions_and_weight() 內部計算時使用，計算完「材積待複核」彙總欄後即丟棄，
# 不會出現在乾淨/隔離資料主表裡（Part 11 起改為濃縮呈現，理由見該函式內部註解）。
MATERIAL_DETAIL_COLUMNS = (
    [f"{c}_原始無法解析_flag" for c in VOLUME_ALL_COLUMNS]
    + [f"{c}_原為負值_flag" for c in VOLUME_ALL_COLUMNS]
    + [f"{c}_原始為0_flag" for c in VOLUME_ALL_COLUMNS]
    + [f"{c}_極端值_flag" for c in VOLUME_ALL_COLUMNS]
    + [f"{c}_共用可疑值_flag" for c in VOLUME_ALL_COLUMNS]
    + [f"{c}_校正來源" for c in VOLUME_ALL_COLUMNS]
    + ["尺寸重量_原始為0_flag", "尺寸重量_極端值_flag", "材積_cm3_極端值_flag"]
)
# 乾淨/隔離資料主表最終保留的材積相關欄位：材積數值＋來源標記＋儲位分類＋彙總複核旗標。
# 材積來源（Part 15）：實測／類別推估／無法計算(缺維度)。儲位分類（Part 16）：貨架／棧板。
MATERIAL_OUTPUT_COLUMNS = ["材積_cm3", "材積來源", "儲位分類", "材積待複核"]

# ============================================================
# 配送地址補齊常數（Part 18，供區域併單使用）
# ============================================================

ADDRESS_COLUMN = "配送地址"
STORE_CODE_COLUMN = "門市代碼"

# 現制縣市（含直轄市／省轄市／縣），用於判斷地址是否已有完整縣市
CURRENT_CITY_RE = re.compile(
    r"^(台北市|臺北市|新北市|桃園市|台中市|臺中市|台南市|臺南市|高雄市|基隆市|"
    r"新竹市|新竹縣|嘉義市|嘉義縣|苗栗縣|彰化縣|南投縣|雲林縣|屏東縣|宜蘭縣|"
    r"花蓮縣|台東縣|臺東縣|澎湖縣|金門縣|連江縣)"
)
# 舊制縣（2010年高雄/台南/台北/台中縣市合併、2014年桃園縣升格）→ 現制直轄市對照
OLD_COUNTY_TO_NEW_CITY = {
    "台北縣": "新北市", "臺北縣": "新北市",
    "高雄縣": "高雄市",
    "台南縣": "台南市", "臺南縣": "台南市",
    "台中縣": "台中市", "臺中縣": "台中市",
    "桃園縣": "桃園市",
}
# 舊制鄉鎮市改制為區時，除了尾碼 市/鎮/鄉→區 的通用規則外，少數鄉鎮同時更名
# （目前資料集中僅發現高雄市那瑪夏區這個案例，如未來遇到其他特殊更名鄉鎮，於此擴充）
SPECIAL_TOWNSHIP_RENAME = {"三民鄉": "那瑪夏區"}
# 比對「舊制縣名」後緊接的鄉鎮市名稱（例："高雄縣岡山鎮..." → 縣後比對出"岡山鎮"）
OLD_TOWNSHIP_RE = re.compile(r"^([^\s\d]{1,3}?)(市|鎮|鄉)")
# 比對地址開頭的「區/鄉/鎮/市」層級行政區（用於判斷是否已有區域，及取出區域文字本身）
DISTRICT_RE = re.compile(r"^([^\s\d]{1,4}?(市|鎮|鄉|區))")
# 少數區名本身第二個字剛好也是 市/鎮/鄉/區 其中之一（例：高雄市「前鎮區」中的「鎮」、
# 桃園市「平鎮區」中的「鎮」、台南市「左鎮區」中的「鎮」），會讓上面 DISTRICT_RE 的
# 非貪婪比對在還沒吃到最後的「區」字之前，提前於「前鎮」/「平鎮」/「左鎮」處誤判為
# 完整區名，漏掉「區」字（Part 22 新增地址回推機制時複核發現，屬既有 Part 18 邏輯
# 的既存問題，非新機制引入）。比對前先攔截這幾個已知特例，其餘一般區名不受影響。
SPECIAL_DISTRICT_NAMES = ["前鎮區", "平鎮區", "左鎮區"]

# 共用可疑值偵測（比照 Part 10 材積清洗 SHARED_VALUE_MIN_DISTINCT_SKU 的手法，但額外加上
# 「單一門市代碼下筆數要夠稀疏」的條件，理由見下方實測發現）：
# 同一組「看似完整」的地址字串，若被過多不同門市代碼「稀疏使用」（每個代碼底下筆數都很少），
# 較可能是系統預設值/測試值，而非任何一個門市的真實地址。
#   實測發現1（真預設值）：「高雄縣橋頭鄉新莊村大學南路1號」被10個門市代碼用過，其中7個代碼
#   「僅出現過2筆、且都在2024-07-05同一天」，符合預設值特徵；但另外3個代碼(5747/5919/6174)
#   各自用了36~87筆、橫跨數個月到一年以上，較可能是這3間門市當時的真實地址（後來才搬遷），
#   不應該被一併排除。
#   實測發現2（誤判案例，已修正）：初版只看「distinct代碼數」不看筆數，曾誤判「高雄市大寮區
#   鳳屏一路107號」為預設值——這組地址被6個門市代碼共用，但每個代碼底下都有15~1040筆、且
#   橫跨整個資料時間範圍(2024-07~2026-05)，明顯是6間真實門市共用同一棟建物地址（例如同一
#   賣場/園區內的不同專櫃），若整組排除會讓這6個代碼原本可信的地址被誤標成待確認。
#   因此改為「只排除同一地址下，筆數<=SHARED_ADDR_SPARSE_MAX_ROWS_PER_CODE 的那些(門市代碼,
#   地址)配對」，而非整組地址全域排除；筆數夠多的配對(如鳳屏一路的6個代碼、大學南路的3個
#   高量代碼)仍視為該代碼的可信版本，正常進入單一/多版本回填邏輯。
SHARED_ADDR_MIN_DISTINCT_CODES = 5
SHARED_ADDR_SPARSE_MAX_ROWS_PER_CODE = 3

# ---- 地址前綴回推（Part 22 新增）----
# 依人工複核發現：先前被標記「查無可信版本、待確認」的地址裡，有不少其實有跡可循，
# 只是省略或簡寫了外層縣市，並非真的無法辨識。新增三種回推規則，套用順序見
# normalize_address_prefix()：

# 1. 常見城市簡寫 → 正式縣市名稱。只收錄實務上常見、不會與其他縣市名稱混淆的簡寫
#    （例如「北市」不會是其他縣市的簡寫，可安全展開；模糊或罕見簡寫不收錄，避免誤判）。
CITY_ABBREVIATIONS = {
    "北市": "台北市", "新北": "新北市", "桃市": "桃園市",
    "中市": "台中市", "南市": "台南市", "高市": "高雄市",
    "竹市": "新竹市", "竹縣": "新竹縣", "苗縣": "苗栗縣",
    "投縣": "南投縣", "雲縣": "雲林縣", "嘉市": "嘉義市",
    "嘉縣": "嘉義縣", "屏縣": "屏東縣", "宜縣": "宜蘭縣",
    "花縣": "花蓮縣", "東縣": "台東縣", "澎縣": "澎湖縣",
    "金縣": "金門縣", "連縣": "連江縣",
}

# 2. 全國唯一（不會與其他縣市同名）的鄉/鎮/縣轄市 → 所屬縣，供地址省略縣名時回推
#    （例如「屏東市」「琉球鄉」全國僅此一個，可安全回推所屬縣；直轄市/省轄市轄下的
#    「區」則刻意不收錄——區名在不同城市間常重複，如中山區、中正區同時存在於台北市、
#    高雄市等多個直轄市，無法安全回推，仍維持人工確認，避免補錯縣市）。
TOWNSHIP_TO_COUNTY = {
    # 新竹縣
    "竹北市": "新竹縣", "竹東鎮": "新竹縣", "新埔鎮": "新竹縣", "關西鎮": "新竹縣",
    "湖口鄉": "新竹縣", "新豐鄉": "新竹縣", "芎林鄉": "新竹縣", "橫山鄉": "新竹縣",
    "北埔鄉": "新竹縣", "寶山鄉": "新竹縣", "峨眉鄉": "新竹縣", "尖石鄉": "新竹縣", "五峰鄉": "新竹縣",
    # 苗栗縣
    "苗栗市": "苗栗縣", "頭份市": "苗栗縣", "苑裡鎮": "苗栗縣", "通霄鎮": "苗栗縣",
    "竹南鎮": "苗栗縣", "後龍鎮": "苗栗縣", "卓蘭鎮": "苗栗縣", "大湖鄉": "苗栗縣",
    "公館鄉": "苗栗縣", "銅鑼鄉": "苗栗縣", "南庄鄉": "苗栗縣", "頭屋鄉": "苗栗縣",
    "三義鄉": "苗栗縣", "西湖鄉": "苗栗縣", "造橋鄉": "苗栗縣", "三灣鄉": "苗栗縣",
    "獅潭鄉": "苗栗縣", "泰安鄉": "苗栗縣",
    # 彰化縣
    "彰化市": "彰化縣", "鹿港鎮": "彰化縣", "和美鎮": "彰化縣", "線西鄉": "彰化縣",
    "伸港鄉": "彰化縣", "福興鄉": "彰化縣", "秀水鄉": "彰化縣", "花壇鄉": "彰化縣",
    "芬園鄉": "彰化縣", "員林市": "彰化縣", "溪湖鎮": "彰化縣", "田中鎮": "彰化縣",
    "大村鄉": "彰化縣", "埔鹽鄉": "彰化縣", "埔心鄉": "彰化縣", "永靖鄉": "彰化縣",
    "社頭鄉": "彰化縣", "二水鄉": "彰化縣", "北斗鎮": "彰化縣", "二林鎮": "彰化縣",
    "田尾鄉": "彰化縣", "埤頭鄉": "彰化縣", "芳苑鄉": "彰化縣", "大城鄉": "彰化縣",
    "竹塘鄉": "彰化縣", "溪州鄉": "彰化縣",
    # 南投縣
    "南投市": "南投縣", "埔里鎮": "南投縣", "草屯鎮": "南投縣", "竹山鎮": "南投縣",
    "集集鎮": "南投縣", "名間鄉": "南投縣", "鹿谷鄉": "南投縣", "中寮鄉": "南投縣",
    "魚池鄉": "南投縣", "國姓鄉": "南投縣", "水里鄉": "南投縣", "信義鄉": "南投縣", "仁愛鄉": "南投縣",
    # 雲林縣
    "斗六市": "雲林縣", "斗南鎮": "雲林縣", "虎尾鎮": "雲林縣", "西螺鎮": "雲林縣",
    "土庫鎮": "雲林縣", "北港鎮": "雲林縣", "古坑鄉": "雲林縣", "大埤鄉": "雲林縣",
    "莿桐鄉": "雲林縣", "林內鄉": "雲林縣", "二崙鄉": "雲林縣", "崙背鄉": "雲林縣",
    "麥寮鄉": "雲林縣", "東勢鄉": "雲林縣", "褒忠鄉": "雲林縣", "台西鄉": "雲林縣",
    "元長鄉": "雲林縣", "四湖鄉": "雲林縣", "口湖鄉": "雲林縣", "水林鄉": "雲林縣",
    # 嘉義縣
    "太保市": "嘉義縣", "朴子市": "嘉義縣", "布袋鎮": "嘉義縣", "大林鎮": "嘉義縣",
    "民雄鄉": "嘉義縣", "溪口鄉": "嘉義縣", "新港鄉": "嘉義縣", "六腳鄉": "嘉義縣",
    "東石鄉": "嘉義縣", "義竹鄉": "嘉義縣", "鹿草鄉": "嘉義縣", "水上鄉": "嘉義縣",
    "中埔鄉": "嘉義縣", "竹崎鄉": "嘉義縣", "梅山鄉": "嘉義縣", "番路鄉": "嘉義縣",
    "大埔鄉": "嘉義縣", "阿里山鄉": "嘉義縣",
    # 屏東縣
    "屏東市": "屏東縣", "潮州鎮": "屏東縣", "東港鎮": "屏東縣", "恆春鎮": "屏東縣",
    "萬丹鄉": "屏東縣", "長治鄉": "屏東縣", "麟洛鄉": "屏東縣", "九如鄉": "屏東縣",
    "里港鄉": "屏東縣", "鹽埔鄉": "屏東縣", "高樹鄉": "屏東縣", "萬巒鄉": "屏東縣",
    "內埔鄉": "屏東縣", "竹田鄉": "屏東縣", "新埤鄉": "屏東縣", "枋寮鄉": "屏東縣",
    "新園鄉": "屏東縣", "崁頂鄉": "屏東縣", "林邊鄉": "屏東縣", "南州鄉": "屏東縣",
    "佳冬鄉": "屏東縣", "琉球鄉": "屏東縣", "車城鄉": "屏東縣", "滿州鄉": "屏東縣",
    "枋山鄉": "屏東縣", "三地門鄉": "屏東縣", "霧台鄉": "屏東縣", "瑪家鄉": "屏東縣",
    "泰武鄉": "屏東縣", "來義鄉": "屏東縣", "春日鄉": "屏東縣", "獅子鄉": "屏東縣", "牡丹鄉": "屏東縣",
    # 宜蘭縣
    "宜蘭市": "宜蘭縣", "羅東鎮": "宜蘭縣", "蘇澳鎮": "宜蘭縣", "頭城鎮": "宜蘭縣",
    "礁溪鄉": "宜蘭縣", "壯圍鄉": "宜蘭縣", "員山鄉": "宜蘭縣", "五結鄉": "宜蘭縣",
    "冬山鄉": "宜蘭縣", "三星鄉": "宜蘭縣", "大同鄉": "宜蘭縣", "南澳鄉": "宜蘭縣",
    # 花蓮縣
    "花蓮市": "花蓮縣", "鳳林鎮": "花蓮縣", "玉里鎮": "花蓮縣", "新城鄉": "花蓮縣",
    "吉安鄉": "花蓮縣", "壽豐鄉": "花蓮縣", "光復鄉": "花蓮縣", "豐濱鄉": "花蓮縣",
    "瑞穗鄉": "花蓮縣", "富里鄉": "花蓮縣", "秀林鄉": "花蓮縣", "萬榮鄉": "花蓮縣", "卓溪鄉": "花蓮縣",
    # 台東縣
    "台東市": "台東縣", "臺東市": "台東縣", "成功鎮": "台東縣", "關山鎮": "台東縣",
    "卑南鄉": "台東縣", "大武鄉": "台東縣", "太麻里鄉": "台東縣", "東河鄉": "台東縣",
    "長濱鄉": "台東縣", "鹿野鄉": "台東縣", "池上鄉": "台東縣", "綠島鄉": "台東縣",
    "延平鄉": "台東縣", "海端鄉": "台東縣", "達仁鄉": "台東縣", "金峰鄉": "台東縣", "蘭嶼鄉": "台東縣",
    # 澎湖縣
    "馬公市": "澎湖縣", "湖西鄉": "澎湖縣", "白沙鄉": "澎湖縣", "西嶼鄉": "澎湖縣",
    "望安鄉": "澎湖縣", "七美鄉": "澎湖縣",
    # 金門縣
    "金城鎮": "金門縣", "金湖鎮": "金門縣", "金沙鎮": "金門縣",
    "金寧鄉": "金門縣", "烈嶼鄉": "金門縣", "烏坵鄉": "金門縣",
    # 連江縣
    "南竿鄉": "連江縣", "北竿鄉": "連江縣", "莒光鄉": "連江縣", "東引鄉": "連江縣",
}
# 比對地址開頭是否為 TOWNSHIP_TO_COUNTY 表中的鄉/鎮/縣轄市名稱。改用「已知名稱清單
# 逐一比對、依長度由長至短優先」而非泛用的「非貪婪正規表示式配市/鎮/鄉結尾」，是因為
# 後者在名稱『中間』剛好也出現市/鎮/鄉其中一字時會提前收尾誤判（Part 22 複核 DISTRICT_RE
# 的「前鎮區」問題時發現同類風險，例如假設性的名稱若中間帶「市」字就可能被誤判，這裡
# 直接用完整已知清單比對可完全避免此類問題，且 TOWNSHIP_TO_COUNTY 是封閉、已知的固定
# 清單，逐一比對成本可忽略）。
_TOWNSHIP_MATCH_RE = re.compile(
    "^(" + "|".join(re.escape(k) for k in sorted(TOWNSHIP_TO_COUNTY, key=len, reverse=True)) + ")"
)

# ---- 舊制縣轄市/鎮/鄉，已於2010(高雄/台南/台北/台中)、2014(桃園)併入直轄市改制為
# 「區」（Part 23 新增）----
# 依人工複核發現：部分地址省略了舊制縣名，直接以「舊制鄉鎮市名稱」開頭（例：「鳳山市
# 海洋一路121號」，省略了「高雄縣」），這種寫法原有的 normalize_old_county_address()
# 無法處理——該函式要求地址必須先出現完整的舊制縣名（如「高雄縣」）才會觸發轉換。
# 新增此表，讓地址開頭直接是舊制鄉鎮市名稱時，也能回推轉換為現制「直轄市＋區」寫法
# （例：鳳山市→高雄市鳳山區）。
# 注意：舊制桃園縣的縣治「桃園市」本身刻意不收錄——現制桃園市（直轄市本身）同名，
# 若地址僅寫「桃園市...」不代表就是指舊桃園市(今桃園區)，可能是現制桃園市底下任一區
# 只是省略了區名，貿然回推有補錯風險，仍交由既有 CURRENT_CITY_RE 判斷為已有縣市、
# 只是缺區，維持人工確認。
OLD_TOWNSHIP_TO_NEW_DISTRICT = {
    # 舊高雄縣（2010年併入高雄市）
    "鳳山市": ("高雄市", "鳳山區"), "岡山鎮": ("高雄市", "岡山區"), "旗山鎮": ("高雄市", "旗山區"),
    "美濃鎮": ("高雄市", "美濃區"), "林園鄉": ("高雄市", "林園區"), "大寮鄉": ("高雄市", "大寮區"),
    "大樹鄉": ("高雄市", "大樹區"), "大社鄉": ("高雄市", "大社區"), "仁武鄉": ("高雄市", "仁武區"),
    "鳥松鄉": ("高雄市", "鳥松區"), "橋頭鄉": ("高雄市", "橋頭區"), "燕巢鄉": ("高雄市", "燕巢區"),
    "田寮鄉": ("高雄市", "田寮區"), "阿蓮鄉": ("高雄市", "阿蓮區"), "路竹鄉": ("高雄市", "路竹區"),
    "湖內鄉": ("高雄市", "湖內區"), "茄萣鄉": ("高雄市", "茄萣區"), "永安鄉": ("高雄市", "永安區"),
    "彌陀鄉": ("高雄市", "彌陀區"), "梓官鄉": ("高雄市", "梓官區"), "六龜鄉": ("高雄市", "六龜區"),
    "甲仙鄉": ("高雄市", "甲仙區"), "杉林鄉": ("高雄市", "杉林區"), "內門鄉": ("高雄市", "內門區"),
    "茂林鄉": ("高雄市", "茂林區"), "桃源鄉": ("高雄市", "桃源區"), "三民鄉": ("高雄市", "那瑪夏區"),
    # 舊台南縣（2010年併入台南市）
    "新營市": ("台南市", "新營區"), "鹽水鎮": ("台南市", "鹽水區"), "白河鎮": ("台南市", "白河區"),
    "柳營鄉": ("台南市", "柳營區"), "後壁鄉": ("台南市", "後壁區"), "東山鄉": ("台南市", "東山區"),
    "麻豆鎮": ("台南市", "麻豆區"), "下營鄉": ("台南市", "下營區"), "六甲鄉": ("台南市", "六甲區"),
    "官田鄉": ("台南市", "官田區"), "大內鄉": ("台南市", "大內區"), "佳里鎮": ("台南市", "佳里區"),
    "學甲鎮": ("台南市", "學甲區"), "西港鄉": ("台南市", "西港區"), "七股鄉": ("台南市", "七股區"),
    "將軍鄉": ("台南市", "將軍區"), "北門鄉": ("台南市", "北門區"), "新化鎮": ("台南市", "新化區"),
    "善化鎮": ("台南市", "善化區"), "新市鄉": ("台南市", "新市區"), "安定鄉": ("台南市", "安定區"),
    "山上鄉": ("台南市", "山上區"), "玉井鄉": ("台南市", "玉井區"), "楠西鄉": ("台南市", "楠西區"),
    "南化鄉": ("台南市", "南化區"), "左鎮鄉": ("台南市", "左鎮區"), "仁德鄉": ("台南市", "仁德區"),
    "歸仁鄉": ("台南市", "歸仁區"), "關廟鄉": ("台南市", "關廟區"), "龍崎鄉": ("台南市", "龍崎區"),
    "永康市": ("台南市", "永康區"),
    # 舊台北縣（2010年升格新北市）
    "板橋市": ("新北市", "板橋區"), "三重市": ("新北市", "三重區"), "中和市": ("新北市", "中和區"),
    "永和市": ("新北市", "永和區"), "新莊市": ("新北市", "新莊區"), "新店市": ("新北市", "新店區"),
    "樹林市": ("新北市", "樹林區"), "鶯歌鎮": ("新北市", "鶯歌區"), "三峽鎮": ("新北市", "三峽區"),
    "淡水鎮": ("新北市", "淡水區"), "汐止市": ("新北市", "汐止區"), "瑞芳鎮": ("新北市", "瑞芳區"),
    "土城市": ("新北市", "土城區"), "蘆洲市": ("新北市", "蘆洲區"), "五股鄉": ("新北市", "五股區"),
    "泰山鄉": ("新北市", "泰山區"), "林口鄉": ("新北市", "林口區"), "深坑鄉": ("新北市", "深坑區"),
    "石碇鄉": ("新北市", "石碇區"), "坪林鄉": ("新北市", "坪林區"), "三芝鄉": ("新北市", "三芝區"),
    "石門鄉": ("新北市", "石門區"), "八里鄉": ("新北市", "八里區"), "平溪鄉": ("新北市", "平溪區"),
    "雙溪鄉": ("新北市", "雙溪區"), "貢寮鄉": ("新北市", "貢寮區"), "金山鄉": ("新北市", "金山區"),
    "萬里鄉": ("新北市", "萬里區"), "烏來鄉": ("新北市", "烏來區"),
    # 舊台中縣（2010年併入台中市）
    "豐原市": ("台中市", "豐原區"), "大里市": ("台中市", "大里區"), "太平市": ("台中市", "太平區"),
    "清水鎮": ("台中市", "清水區"), "沙鹿鎮": ("台中市", "沙鹿區"), "梧棲鎮": ("台中市", "梧棲區"),
    "大甲鎮": ("台中市", "大甲區"), "東勢鎮": ("台中市", "東勢區"), "后里鄉": ("台中市", "后里區"),
    "神岡鄉": ("台中市", "神岡區"), "潭子鄉": ("台中市", "潭子區"), "大雅鄉": ("台中市", "大雅區"),
    "新社鄉": ("台中市", "新社區"), "石岡鄉": ("台中市", "石岡區"), "外埔鄉": ("台中市", "外埔區"),
    "大安鄉": ("台中市", "大安區"), "烏日鄉": ("台中市", "烏日區"), "大肚鄉": ("台中市", "大肚區"),
    "龍井鄉": ("台中市", "龍井區"), "霧峰鄉": ("台中市", "霧峰區"), "和平鄉": ("台中市", "和平區"),
    # 舊桃園縣（2014年升格桃園市；縣治「桃園市」本身不收錄，理由見上方說明）
    "中壢市": ("桃園市", "中壢區"), "平鎮市": ("桃園市", "平鎮區"), "楊梅市": ("桃園市", "楊梅區"),
    "大溪鎮": ("桃園市", "大溪區"), "蘆竹鄉": ("桃園市", "蘆竹區"), "大園鄉": ("桃園市", "大園區"),
    "龜山鄉": ("桃園市", "龜山區"), "八德市": ("桃園市", "八德區"), "龍潭鄉": ("桃園市", "龍潭區"),
    "新屋鄉": ("桃園市", "新屋區"), "觀音鄉": ("桃園市", "觀音區"), "復興鄉": ("桃園市", "復興區"),
}
# 比對地址開頭是否為 OLD_TOWNSHIP_TO_NEW_DISTRICT 表中的舊制鄉鎮市名稱，原理與
# _TOWNSHIP_MATCH_RE 相同：改用已知名稱清單比對，避免「新市鄉」「左鎮鄉」「平鎮市」
# 這類名稱中間剛好也含市/鎮字元的個案被非貪婪正規表示式提前誤判收尾。
_OLD_TOWNSHIP_MATCH_RE = re.compile(
    "^(" + "|".join(re.escape(k) for k in sorted(OLD_TOWNSHIP_TO_NEW_DISTRICT, key=len, reverse=True)) + ")"
)

# ---- 舊制縣名常見簡寫（Part 23 新增）----
# 供地址開頭直接寫簡寫舊制縣名時（例：「高縣茄萣鄉...」），展開為完整舊制縣名，
# 再交由既有的 normalize_old_county_address() 轉換為現制寫法。
OLD_COUNTY_ABBREVIATIONS = {
    "北縣": "台北縣", "高縣": "高雄縣", "南縣": "台南縣", "中縣": "台中縣", "桃縣": "桃園縣",
}

# 3. 開頭誤帶郵遞區號（3~6碼數字，含新式6碼與舊式3碼）：去除後才進行縣市比對。
#    只在去除後的剩餘文字看起來像可辨識地址開頭時才動作，避免誤傷本身就以數字開頭
#    的門牌（實務上台灣地址門牌號一律接在路名之後，不會出現在地址最前面，此風險極低，
#    但仍加上這層檢查以策安全）。
POSTAL_CODE_PREFIX_RE = re.compile(r"^\(?\d{3,6}\)?[\s,，、\-]*")

# 配送地址補齊新增的欄位（供隔離資料輸出時一併帶出，乾淨資料因不限定欄位清單會自動包含）。
# 「配送區域」（Part 21）＝縣市＋區域（例：新北市內湖區），不含路名門牌，供區域併單策略
# 直接分組使用，不需要自行從完整地址解析。
# Part 25（離線路名回推）新增一欄，見 resolve_addresses_offline()：
#   配送地址_離線比對 ：離線比對標籤（未載入參考資料／路名回推補齊／路名相符(已驗證)／未比對）
ADDRESS_EXTRA_COLUMNS = [
    "配送地址_補值來源", "配送地址_待確認_flag", "配送區域", "配送地址_離線比對",
]

# 商品類別對應表：69 碼固定版本
# (商品類別ID, 商品類別代碼, 商品類別名稱, 來源)
# 前 38 碼為原始 P&G 對照表（其中 5 碼由 3M 商品資料判讀補上正式名稱）
# 中間 20 碼、後 11 碼為依商品名稱關鍵字人工判讀新增
CATEGORY_TABLE = [
    (11, "OCM", "手動牙刷", "原始P&G對照表"),
    (12, "F&HC", "織品及家用品", "原始P&G對照表"),
    (13, "B&R", "刮鬍刀", "原始P&G對照表"),
    (14, "FEM", "女性用品", "原始P&G對照表"),
    (15, "BABY", "嬰兒用品", "原始P&G對照表"),
    (16, "OCP", "電動牙刷", "原始P&G對照表"),
    (17, "HAIR", "髮類", "原始P&G對照表"),
    (18, "SNCK", "品客", "原始P&G對照表"),
    (19, "SKIN", "歐蕾", "原始P&G對照表"),
    (20, "PCC", "香皂", "原始P&G對照表"),
    (21, "OTHER", "其他", "原始P&G對照表"),
    (22, "BATTERY", "電池", "原始P&G對照表"),
    (23, "APPL", "百靈", "原始P&G對照表"),
    (24, "8BK4", "8BK4醫療文具剪刀", "原始P&G對照表"),
    (25, "8AA4", "8AA4家用品", "原始P&G對照表"),
    (26, "009", "吸乳器009", "原始P&G對照表"),
    (27, "008", "孕產婦和哺乳服裝", "原始P&G對照表"),
    (28, "007", "母乳收集及配件", "原始P&G對照表"),
    (29, "006", "乳頭護理系列", "原始P&G對照表"),
    (31, "005", "吸乳器005", "原始P&G對照表"),
    (32, "004", "鋁箔", "原始P&G對照表"),
    (33, "003", "保鮮膜", "原始P&G對照表"),
    (34, "002", "愛樂購", "原始P&G對照表"),
    (35, "APP", "衛生紙", "原始P&G對照表"),
    (36, "8DA4", "8DA4 掛鉤DIY膠", "原始P&G對照表"),
    (40, "WN", "WN", "原始對照表(查無訂單資料，無法推估)"),
    (41, "HESS", "HESS", "原始對照表(查無訂單資料，無法推估)"),
    (42, "O-BW", "O-BW", "原始對照表(查無訂單資料，無法推估)"),
    (43, "BASF", "BASF", "原始對照表(查無訂單資料，無法推估)"),
    (44, "001", "001", "原始對照表(查無訂單資料，無法推估)"),
    (45, "8CB4", "居家清潔劑(3M/魔利)", "原始對照表(名稱經商品資料判讀更新)"),
    (46, "8EE5", "工作手套(3M DIY)", "原始對照表(名稱經商品資料判讀更新)"),
    (47, "8FF5", "8FF5", "原始對照表(查無訂單資料，無法推估)"),
    (48, "85C4", "膠帶及封箱用品(3M)", "原始對照表(名稱經商品資料判讀更新)"),
    (49, "FG01", "FG01", "原始對照表(查無訂單資料，無法推估)"),
    (50, "BL1", "BL1", "原始對照表(查無訂單資料，無法推估)"),
    (55, "8HH5", "汽車美容清潔用品(3M)", "原始對照表(名稱經商品資料判讀更新)"),
    (56, "010", "益生菌保健食品", "原始對照表(名稱經商品資料判讀更新)"),
    (57, "TISSUE", "衛生紙及紙品(舒潔/可麗舒)", "新增-依商品名稱推估"),
    (58, "DIAPER", "紙尿褲/褲型尿布(好奇)", "新增-依商品名稱推估"),
    (59, "FEMCARE", "女性衛生棉/護墊/成人褲(靠得住)", "新增-依商品名稱推估"),
    (60, "KC-GIFT", "品牌玩具/週邊贈品(好奇/靠得住)", "新增-依商品名稱推估"),
    (61, "KCP-WIPE", "工業擦拭布及防護用品(金百利專業)", "新增-依商品名稱推估"),
    (62, "WETWIPE", "濕巾/濕式衛生紙(好奇/舒潔)", "新增-依商品名稱推估"),
    (64, "COSMETIC", "美妝保養品(Dr.Cink等)", "新增-依商品名稱推估"),
    (65, "OTC-DRUG", "指示藥品/成藥", "新增-依商品名稱推估"),
    (66, "ORAL-HC", "口腔清潔及居家清潔用品", "新增-依商品名稱推估"),
    (67, "SUPPLEMENT", "保健食品", "新增-依商品名稱推估"),
    (68, "BABYCARE2", "嬰幼兒沐浴保養品(非P&G品牌)", "新增-依商品名稱推估"),
    (69, "MEDSUPPLY", "醫療耗材/傷口護理用品", "新增-依商品名稱推估"),
    (70, "MEDNUTRI", "特殊營養配方食品", "新增-依商品名稱推估"),
    (71, "SNACK", "零食/休閒食品", "新增-依商品名稱推估"),
    (72, "KITCHEN", "廚房生活小物", "新增-依商品名稱推估"),
    (74, "PET", "寵物用品", "新增-依商品名稱推估"),
    (75, "FEMCARE2", "女性衛生棉/成人紙尿褲(多品牌混合)", "新增-依商品名稱推估"),
    (77, "GIFT", "促銷贈品(多品牌)", "新增-依商品名稱推估"),
    (81, "TEST-HH", "日用品(疑似測試資料)", "新增-依商品名稱推估"),
    (82, "TEST-SNK", "零食(疑似測試資料)", "新增-依商品名稱推估"),
    (83, "SEAFOOD-MEAT", "生鮮/凍藏食材批發(海鮮.肉品.乾貨)", "新增-依商品名稱推估(隔離資料重分類)"),
    (84, "FOOD-WHOLESALE", "食品雜貨/冷凍調理食品(晨家等品牌)", "新增-依商品名稱推估(隔離資料重分類)"),
    (85, "MUCHENSHI", "居家清潔抹布/除塵紙(無塵氏)", "新增-依商品名稱推估(隔離資料重分類)"),
    (86, "WETWIPE-OTHER", "濕巾/化妝棉(其他品牌)", "新增-依商品名稱推估(隔離資料重分類)"),
    (87, "APPLIANCE-COLOR", "小家電/生活家電(色碼系列)", "新增-依商品名稱推估(隔離資料重分類)"),
    (88, "LAUNDRY", "洗衣球/洗衣用品", "新增-依商品名稱推估(隔離資料重分類)"),
    (89, "STATIONERY", "文具/曆本禮品(讀曆書店)", "新增-依商品名稱推估(隔離資料重分類)"),
    (90, "PACKAGING", "包材/寄件紙箱/提袋", "新增-依商品名稱推估(隔離資料重分類)"),
    (91, "GIFTBOX-FOOD", "年節禮盒/年菜", "新增-依商品名稱推估(隔離資料重分類)"),
    (92, "MISC-GADGET", "小家電/生活雜貨(其他)", "新增-依商品名稱推估(隔離資料重分類)"),
    (93, "ADULT-INCONT", "成人紙尿褲/看護褲(自由適)", "新增-依商品名稱推估(隔離資料重分類)"),
]
VALID_CATEGORY_IDS = {row[0] for row in CATEGORY_TABLE}
CATEGORY_NAME_BY_ID = {row[0]: row[2] for row in CATEGORY_TABLE}

# 隔離資料重分類關鍵字規則（依優先序，第一個命中者生效）
# -1 = 疑似測試資料佔位字樣，不分類、僅標記
RECLASSIFY_RULES = [
    (r"測試商品", -1),
    (r"好奇.*(磁性板|收納箱|泡泡相機|滾球|轉轉樂|嚕嚕車|七巧板|電子琴畫板|玩具)", 60),
    (r"好奇.*(濕巾|純水)", 62),
    (r"好奇", 58),
    (r"舒潔.*(濕式|濕巾)|Kleenex.*(濕式|濕巾|卸妝)", 62),
    (r"舒潔|可麗雅|可立雅|可麗舒|Kleenex|OEM.*衛生紙|家樂福.*衛生紙|統一超商.*衛生紙", 57),
    (r"靠得住|Kotex", 59),
    (r"自由適", 93),
    (r"三面電動牙刷", 16),
    (r"丁晴手套|防護衣|工業用紙|Scott|WypAll|AQUARIUS", 61),
    (r"JIGOTT|theSAEM|Prreti|面膜|精華[液露]|眼霜|眼膜|面霜|唇膜|唇線筆|遮瑕|去角質|安瓶|燕窩禮盒|舒敏.*潔膚|護手霜|沐浴露|積雪草|黃金膠原|蘆薈舒緩保濕凝膠", 64),
    (r"寶兒樂|寵物", 74),
    (r"洗衣膠球|洗衣球|蘭諾衣物芳香豆|去漬液", 88),
    (r"^KA[-\d]|^RM-\d|^KK[-\d]|^AD-\d|^\d{5}-\d{2}(芥末黃|淺灰|火山灰|湖藍|啡紅|翡翠綠|櫻花粉|灰綠)|隨行果汁機", 87),
    (r"讀曆|拾光造字本|超能曆|種子曆", 89),
    (r"寄件紙箱|不織布帶扣提袋|紙袋", 90),
    (r"無塵氏", 85),
    (r"拭拭樂|悅拭美|悅仕美|惜福品|比得兔|酷洛米.*方巾|涼感巾|涼感毛巾|化妝棉|純水.*方巾|卸妝濕巾|迷你濕巾", 86),
    (r"晨家|呷米|Kai康普茶|米啤酒|澎湖啤酒|紅點啤酒|高粱|蘭陽食品|有機紅茶豆奶", 84),
    (r"^[一-龥]-", 84),
    (r"Litheli|FunPeak|Washwow|Byee|Tia-One|格外美|不落枕|藍光手持式消毒噴霧槍|Ｋ５ＰＲＯ|折疊蛋捲桌|香氛蠟燭|香氛擴香瓶", 92),
    (r"禮盒|年菜|佛跳牆", 91),
    (r"豌豆|蠶豆|香脆.*豆", 71),
    (r"蝦|魚|蟹|蟳|貝|軟絲|小卷|魷魚|章魚|干貝|鮑|扇貝|海蜇|海參|牡蠣|生蠔|文蛤|鱈|鮭|鯛|鯖|秋刀|旗魚|鰹魚|鰻|龍蝦|紅條|紅板|白昌|白足|白仁|大頭|二頭|老A|大A|市足|市肉|市身|太陽|角蝦|飛魚卵|鱙魚|虱目魚|牛|豬|雞|肉品|腱|五花|骨腿|大腸頭|條肉|棒棒腿|三節翅|沙朗|炭烤|德國豬腳|清腿|前腿心|板腱|帝王蟹|珠貝|海麻仁|海瓜子|香菇|菇|玉米|洋甘菊|冬瓜|地瓜|米菓|米果|散裝|花枝|鳴戶卷|姬路城|錦絲|和尚頭|石班|透抽|生片|熟卷|盒[卷捲]|英哥|大武|中尖|中紅|大紅|小花枝|中母市|母市|公市|中公|大公|白秋|南尤|紫御|培根|火鍋料|昆布|裙帶|甜不辣|土托|龍肉|預打泥|抽[\s(#\d]|A抽|特大抽|中抽|中大抽|扁雪片|竹筍|杏仁|枸杞|黑木耳|藜麥|海帶|烤麩|丁香|蓮藕|草猴|白猴|海鮮筒|海鮮卷|金鑽酥|燒賣|湯包|煎餃|手工肉羹|華昇|金線|印度.*漿|MEENA|NAIK|SeaStar|Yashaswi|OCEAN|UTTAM|SILVER|ODYSSEY|板燒|玉子|豆腐燒|蛋皮|佃煮|羔羊|冷凍腿肉|白米|東門|烏龍麵|青花椒鹽|團膳|明月冰卷|南瓜|圓芝士豆腐|小石支|芝心包|市.*紅市|^\(\d+[-/]|^\(\d+\)|K\)\s*\d", 83),
    (r"零食|軟糖|爆米花|餅乾|喉糖|QQ糖", 71),
]
_COMPILED_RECLASSIFY_RULES = [(re.compile(p), cid) for p, cid in RECLASSIFY_RULES]

PICK_DATE_MIN = pd.Timestamp("2024-01-01")
PICK_DATE_MAX = pd.Timestamp("2026-12-31")


def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


# Part 14：耐大檔的 Excel 讀取。實測本資料集的「訂單資料」工作表解壓後是
# 約 700MB 的單行 XML，新版 openpyxl（3.1.x）的串流解析器會在中途丟出
# xml.etree.ElementTree.ParseError: mismatched tag 而整個讀取失敗（壓縮檔本身完好，
# 純粹是解析器對超大單行 XML 的限制）。改為優先用 calamine（Rust 實作、耐大檔又快），
# 未安裝或讀取失敗時才退回 openpyxl，確保這支腳本可以直接執行、不必外掛包裝器。
#   安裝：pip install python-calamine
def read_excel_robust(path: str, sheet: str) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name=sheet, dtype=object, engine="calamine")
    except Exception as e:
        log(f"⚠️ calamine 讀取失敗或未安裝（{e}），改用 openpyxl 重試"
            f"（若輸入檔的工作表非常大，openpyxl 可能無法解析，建議 pip install python-calamine）")
        return pd.read_excel(path, sheet_name=sheet, dtype=object, engine="openpyxl")


# ============================================================
# 工具函式
# ============================================================

def is_null_token(x) -> bool:
    """判斷是否為『真空值』或 NULL 類字樣（不分大小寫，去除前後空白後比對）"""
    if pd.isna(x):
        return True
    s = str(x).strip()
    return s.upper() in NULL_TOKENS


def normalize_null_series(s: pd.Series) -> pd.Series:
    return s.map(lambda x: np.nan if is_null_token(x) else x)


def trim_series(s: pd.Series) -> pd.Series:
    return s.map(lambda x: x.strip() if isinstance(x, str) else x)


def to_int_display(x):
    """float 15.0 -> 15（整數顯示），缺值維持缺值"""
    if pd.isna(x):
        return np.nan
    try:
        return int(round(float(x)))
    except (ValueError, TypeError):
        return x


def parse_flex_date(x):
    """指定到貨日期／有效日期：文字字串轉日期，NaT 保留，解析失敗也保留原值"""
    if pd.isna(x):
        return np.nan
    if isinstance(x, (pd.Timestamp, datetime, date)):
        return pd.Timestamp(x)
    s = str(x).strip()
    if s == "":
        return np.nan
    try:
        return pd.to_datetime(s)
    except (ValueError, TypeError):
        return x  # 解析失敗，保留原字串（依歷史資料，此腳本涵蓋之格式解析成功率100%）


def parse_flex_date_series(col):
    """parse_flex_date 的整欄向量化版本：百萬列規模時，逐格呼叫 parse_flex_date()
    （其內部逐格呼叫 pd.to_datetime(單一字串)）實測極慢（100萬列一欄要跑數分鐘），改成
    先對整欄一次性做向量化解析（pandas C 層一次掃過，format="mixed" 讓它逐元素自行判斷
    格式、不強迫整欄套同一種格式，效果等同逐格呼叫 dateutil 解析），只有極少數向量化解析
    失敗（NaT）的格子，才退回原本逐格版 parse_flex_date() 求出正確答案（可能是真的 NaT，
    也可能是保留原字串，兩者都有可能、不能用「NaT 就等於失敗」去猜，所以直接呼叫逐格版
    取得唯一正確答案，只是只對這極少數格子做，不影響整體效能）。已用大量邊界案例（含
    NaN／NaN-like 字串／混合日期格式／已是日期型別／解析失敗字串）驗證與逐格版結果
    完全一致。"""
    s = col.copy()
    na_mask = s.isna()
    dt_mask = s.map(lambda v: isinstance(v, (pd.Timestamp, datetime, date))) & ~na_mask
    conv_mask = ~na_mask & ~dt_mask

    result = pd.Series(index=s.index, dtype=object)
    result[na_mask] = np.nan
    if dt_mask.any():
        result[dt_mask] = s[dt_mask].map(pd.Timestamp)

    if conv_mask.any():
        raw = s[conv_mask].astype(str).str.strip()
        empty_mask = raw == ""
        result.loc[raw.index[empty_mask]] = np.nan
        nonempty = raw[~empty_mask]
        if len(nonempty):
            parsed = pd.to_datetime(nonempty, errors="coerce", format="mixed")
            ok_mask = parsed.notna()
            result.loc[parsed.index[ok_mask]] = parsed[ok_mask]
            fail_idx = parsed.index[~ok_mask]
            for idx in fail_idx:
                result.loc[idx] = parse_flex_date(s.loc[idx])
    return result


def deterministic_mode(values):
    """決定性眾數：依出現次數由多到少，次數相同則依字串排序由小到大，取第一個。
    避免 pandas mode() 在極端情況下的非決定性（例如 tie）。"""
    counts = {}
    for v in values:
        if pd.isna(v):
            continue
        counts[v] = counts.get(v, 0) + 1
    if not counts:
        return np.nan
    return sorted(counts.items(), key=lambda kv: (-kv[1], str(kv[0])))[0][0]


def parse_shipno_date(shipno):
    """由系統出貨單號解析出貨日期：OU + 8碼日期(YYYYMMDD) + 6碼序號"""
    if not isinstance(shipno, str) or not SHIPNO_PATTERN.match(shipno):
        return np.nan
    ymd = shipno[2:10]
    try:
        return pd.Timestamp(year=int(ymd[0:4]), month=int(ymd[4:6]), day=int(ymd[6:8]))
    except (ValueError, TypeError):
        return np.nan


def classify_by_keyword(name):
    """回傳 (category_id or None, matched_rule_index or None)。None 代表無任何規則命中。"""
    if not isinstance(name, str):
        return None
    for pattern, cat_id in _COMPILED_RECLASSIFY_RULES:
        if pattern.search(name):
            return cat_id
    return None


def is_nondescriptive_name(name) -> bool:
    """Part 17：判斷商品名稱是否幾乎沒有可辨識品名文字（純規格代號/數字符號，例如「(60) 10/15」）。
    缺值名稱不在此判定（已由規則4處理），只針對『有字串但沒有意義文字』者。"""
    if not isinstance(name, str):
        return False
    s = name.strip()
    if s == "":
        return False
    return len(_MEANINGFUL_CHAR_RE.findall(s)) < NAME_MIN_MEANINGFUL_CHARS


def is_test_placeholder_name(name) -> bool:
    """判斷商品名稱是否符合測試/範例資料的命名模式（Part 10 新增，供商品材積主檔排除用）。
    範圍比 RECLASSIFY_RULES 裡的「測試商品」規則更廣，涵蓋更多常見的假資料命名慣例。"""
    if not isinstance(name, str):
        return False
    return any(p.search(name) for p in _COMPILED_TEST_DATA_PATTERNS)


def normalize_old_county_address(addr: str) -> str:
    """地址開頭若為舊制縣名（高雄縣／台南縣／台北縣／台中縣／桃園縣），轉換為現制
    直轄市名稱，並將緊接的鄉/鎮/市尾碼一併轉換為「區」（Part 18，供區域併單分組一致）。
    非上述舊制縣開頭的地址，原樣傳回不做任何更動。"""
    if not isinstance(addr, str) or addr == "":
        return addr
    for old_county, new_city in OLD_COUNTY_TO_NEW_CITY.items():
        if addr.startswith(old_county):
            rest = addr[len(old_county):]
            m = OLD_TOWNSHIP_RE.match(rest)
            if not m:
                # 縣名後接不到可辨識的鄉鎮市名稱，只轉換縣名本身，鄉鎮市部分保留原樣待人工複核
                return new_city + rest
            township = m.group(0)
            new_district = SPECIAL_TOWNSHIP_RENAME.get(township, m.group(1) + "區")
            return new_city + new_district + rest[len(township):]
    return addr


def expand_city_abbreviation(addr: str) -> str:
    """地址開頭若為常見城市簡寫（例：北市→台北市），展開為正式縣市名稱（Part 22）。
    非上述簡寫開頭的地址，原樣傳回不做任何更動。地址若已是完整正式縣市名稱則直接
    略過——例如「新北」是「新北市」的簡寫，但同時也是「新北市」字面上的前綴，若不
    先排除已完整的情況，「新北市...」會被誤展開成「新北市市...」。"""
    if not isinstance(addr, str) or addr == "":
        return addr
    if CURRENT_CITY_RE.match(addr):
        return addr
    for abbr, full in CITY_ABBREVIATIONS.items():
        if addr.startswith(abbr):
            return full + addr[len(abbr):]
    return addr


def infer_county_from_township(addr: str) -> str:
    """地址開頭若省略縣名、直接是全國唯一（無同名歧義）的鄉/鎮/縣轄市名稱
    （例：屏東市、琉球鄉），依 TOWNSHIP_TO_COUNTY 回推所屬縣、補回縣名（Part 22）。
    地址已有縣市開頭，或開頭不是表中任何鄉鎮市名稱者，原樣傳回不做更動。"""
    if not isinstance(addr, str) or addr == "":
        return addr
    if CURRENT_CITY_RE.match(addr):
        return addr
    m = _TOWNSHIP_MATCH_RE.match(addr)
    if not m:
        return addr
    return TOWNSHIP_TO_COUNTY[m.group(1)] + addr


def expand_old_county_abbreviation(addr: str) -> str:
    """地址開頭若為舊制縣名簡寫（例：「高縣」→高雄縣），展開為完整舊制縣名，
    供後續 normalize_old_county_address() 接手轉換為現制寫法（Part 23）。"""
    if not isinstance(addr, str) or addr == "":
        return addr
    if CURRENT_CITY_RE.match(addr):
        return addr
    for abbr, full in OLD_COUNTY_ABBREVIATIONS.items():
        if addr.startswith(abbr):
            return full + addr[len(abbr):]
    return addr


def infer_new_district_from_old_township(addr: str) -> str:
    """地址開頭若省略了舊制縣名，直接是「已併入直轄市的舊制鄉鎮市名稱」
    （例：「鳳山市」，省略了「高雄縣」），依 OLD_TOWNSHIP_TO_NEW_DISTRICT 回推轉換
    為現制「直轄市＋區」寫法（例：鳳山市→高雄市鳳山區）（Part 23）。地址已有現制
    縣市開頭，或開頭不是表中任何舊制鄉鎮市名稱者，原樣傳回不做更動。"""
    if not isinstance(addr, str) or addr == "":
        return addr
    if CURRENT_CITY_RE.match(addr):
        return addr
    m = _OLD_TOWNSHIP_MATCH_RE.match(addr)
    if not m:
        return addr
    new_city, new_district = OLD_TOWNSHIP_TO_NEW_DISTRICT[m.group(1)]
    return new_city + new_district + addr[m.end():]


def _looks_like_recognizable_address_start(s: str) -> bool:
    """判斷字串開頭是否已經『看起來像』地址開頭（縣市全名/簡寫/舊制縣名(簡寫)/
    唯一鄉鎮市名稱/舊制已併入鄉鎮市名稱之一），供 strip_postal_code_prefix() 確認
    去除疑似郵遞區號後的剩餘文字合理，避免誤傷。"""
    if not isinstance(s, str) or s == "":
        return False
    if CURRENT_CITY_RE.match(s):
        return True
    if any(s.startswith(old) for old in OLD_COUNTY_TO_NEW_CITY):
        return True
    if any(s.startswith(abbr) for abbr in CITY_ABBREVIATIONS):
        return True
    if any(s.startswith(abbr) for abbr in OLD_COUNTY_ABBREVIATIONS):
        return True
    if _TOWNSHIP_MATCH_RE.match(s):
        return True
    return bool(_OLD_TOWNSHIP_MATCH_RE.match(s))


def strip_postal_code_prefix(addr: str) -> str:
    """地址開頭若誤帶郵遞區號（3~6碼數字，含新式6碼與舊式3碼，可能帶括號/逗號等
    分隔符），去除後才進行縣市比對（Part 22）。只有在去除後的剩餘文字看起來像可
    辨識地址開頭時才動作，避免誤傷本身就以數字開頭、剛好不是地址的內容。"""
    if not isinstance(addr, str) or addr == "":
        return addr
    m = POSTAL_CODE_PREFIX_RE.match(addr)
    if not m or m.end() == 0:
        return addr
    rest = addr[m.end():]
    return rest if _looks_like_recognizable_address_start(rest) else addr


def normalize_address_prefix(addr: str) -> str:
    """地址前綴回推主流程（Part 22／Part 23）：依序處理幾種先前會被誤判為『查無
    可信版本、待確認』、但其實有跡可循的寫法，處理完再交由既有的新舊制行政區更名
    正規化：
      1. 去除誤帶的郵遞區號前綴。
      2. 展開常見「現制」城市簡寫（北市→台北市）。
      3. 展開常見「舊制」縣名簡寫（高縣→高雄縣），供下一步接手。
      4. 新舊制行政區更名正規化（原有 Part 18 邏輯，例：高雄縣→高雄市，含完整
         舊制縣名+鄉鎮市寫法）。
      5. 省略舊制縣名、直接是已併入直轄市的舊制鄉鎮市名稱者，回推轉換為現制
         「直轄市＋區」寫法（例：鳳山市→高雄市鳳山區）。
      6. 省略現制縣名、只寫全國唯一鄉鎮市名稱者，回推補上所屬縣
         （例：屏東市→屏東縣屏東市）。
    """
    addr = strip_postal_code_prefix(addr)
    addr = expand_city_abbreviation(addr)
    addr = expand_old_county_abbreviation(addr)
    addr = normalize_old_county_address(addr)
    addr = infer_new_district_from_old_township(addr)
    addr = infer_county_from_township(addr)
    return addr


def split_city_district_rest(addr: str):
    """將（已正規化的）地址拆成 (縣市或None, 區域或None, 其餘路名門牌文字)。
    僅依現制縣市/區域文字表面比對，不驗證路名門牌是否真實存在（Part 18）。"""
    if not isinstance(addr, str) or addr == "":
        return None, None, addr if isinstance(addr, str) else ""
    m_city = CURRENT_CITY_RE.match(addr)
    if m_city:
        city = m_city.group(1)
        rest = addr[len(city):]
    else:
        city = None
        rest = addr
    special = next((d for d in SPECIAL_DISTRICT_NAMES if rest.startswith(d)), None)
    if special:
        district = special
        rest = rest[len(special):]
    else:
        m_dist = DISTRICT_RE.match(rest)
        if m_dist:
            district = m_dist.group(1)
            rest = rest[len(district):]
        else:
            district = None
    return city, district, rest


# ============================================================
# 1. 讀取與基礎正規化（清洗文件.md Part 2 §0、§1）
# ============================================================

def load_and_normalize(input_path: str):
    log(f"讀取原始檔案：{input_path}")
    df = read_excel_robust(input_path, "訂單資料")
    df.columns = [str(c).strip() for c in df.columns]  # 修正表頭前後空白（如" 門市代碼"）
    df = df[ORDER_COLUMNS].copy()
    df.insert(0, "_原始列號", range(2, len(df) + 2))  # 對齊 Excel 列號（表頭為第1列）
    original_row_count = len(df)
    log(f"原始資料列數（不含表頭）：{original_row_count}")

    log("步驟1：NULL字樣→真空值正規化（門市訂單號）")
    for col in NULL_NORMALIZE_COLUMNS:
        df[col] = normalize_null_series(df[col])

    log("步驟2：去除前後空白（商品編號／商品名稱／單位）")
    for col in TRIM_COLUMNS:
        df[col] = trim_series(df[col])

    log("步驟3：商品ID／商品類別 float顯示 → 整數顯示")
    df["商品ID"] = df["商品ID"].map(to_int_display)
    df["商品類別"] = df["商品類別"].map(to_int_display)

    log("步驟5：訂購數量 轉為數值型別（供規則判斷使用，缺值/非數字維持原樣）")
    df["訂購數量_數值"] = pd.to_numeric(df["訂購數量"], errors="coerce")

    return df, original_row_count


# ============================================================
# 2. 八條優先序規則：分流「乾淨資料」與「隔離資料」
#    （清洗文件.md Part 2 §2、Part 6 §2-7）
# ============================================================

def apply_isolation_rules(df: pd.DataFrame, raw_df_for_dup_check: pd.DataFrame):
    n = len(df)
    reason = pd.Series([None] * n, index=df.index, dtype=object)
    change_log = []  # (原始列號, 欄位, 規則, 原始值, 清理後值)

    # ---- 規則1：系統出貨單號格式異常 ----
    valid_shipno = df["系統出貨單號"].map(
        lambda x: isinstance(x, str) and bool(SHIPNO_PATTERN.match(x))
    )
    rule1_hit = ~valid_shipno
    log(f"規則1 系統出貨單號格式異常：命中 {int(rule1_hit.sum())} 筆")
    reason[rule1_hit & reason.isna()] = "系統出貨單號格式異常"

    # ---- 規則2：完全重複列（以「清理前的原始值」14欄完全相同為準，僅保留第一筆）----
    dup_mask = raw_df_for_dup_check.duplicated(subset=ORDER_COLUMNS, keep="first")
    log(f"規則2 完全重複列（原始值）：命中 {int(dup_mask.sum())} 筆")
    reason[dup_mask.values & reason.isna()] = "完全重複列"

    # ---- 規則3：商品ID缺失 ----
    id_missing = df["商品ID"].isna()
    log(f"規則3 商品ID缺失：命中 {int(id_missing.sum())} 筆")
    reason[id_missing & reason.isna()] = "商品ID缺失"

    # ---- 商品名稱補正（依商品ID多數值補齊，缺值時查無法補齊者留待規則4判斷）----
    name_mode_by_id = (
        df.loc[df["商品名稱"].notna() & df["商品ID"].notna()]
        .groupby("商品ID")["商品名稱"]
        .apply(deterministic_mode)
    )
    name_missing = df["商品名稱"].isna()
    fillable_name = name_missing & df["商品ID"].map(lambda x: x in name_mode_by_id.index if pd.notna(x) else False)
    for idx in df.index[fillable_name]:
        pid = df.at[idx, "商品ID"]
        new_name = name_mode_by_id.get(pid, np.nan)
        if pd.notna(new_name):
            change_log.append((df.at[idx, "_原始列號"], "商品名稱", "缺值時依商品ID多數值補齊",
                                df.at[idx, "商品名稱"], new_name))
            df.at[idx, "商品名稱"] = new_name
    name_still_missing = df["商品名稱"].isna()
    log(f"商品名稱依ID多數值補齊：成功 {int((fillable_name & ~name_still_missing).sum())} 筆")

    # ---- 規則4：商品名稱缺值且無法補齊 ----
    rule4_hit = name_still_missing
    log(f"規則4 商品名稱缺值且無法補齊：命中 {int(rule4_hit.sum())} 筆")
    reason[rule4_hit.values & reason.isna()] = "商品名稱缺值且無法補齊"

    # ---- 商品類別補正（兩層：先依商品ID多數值，再依商品名稱多數值）----
    cat_mode_by_id = (
        df.loc[df["商品類別"].notna() & df["商品ID"].notna()]
        .groupby("商品ID")["商品類別"]
        .apply(deterministic_mode)
    )
    cat_missing = df["商品類別"].isna()
    fillable_cat_by_id = cat_missing & df["商品ID"].map(lambda x: x in cat_mode_by_id.index if pd.notna(x) else False)
    filled_by_id = 0
    for idx in df.index[fillable_cat_by_id]:
        pid = df.at[idx, "商品ID"]
        new_cat = cat_mode_by_id.get(pid, np.nan)
        if pd.notna(new_cat):
            change_log.append((df.at[idx, "_原始列號"], "商品類別", "缺值時依商品ID多數值補齊",
                                df.at[idx, "商品類別"], new_cat))
            df.at[idx, "商品類別"] = new_cat
            filled_by_id += 1
    log(f"商品類別第一層(依ID)補齊：成功 {filled_by_id} 筆")

    cat_mode_by_name = (
        df.loc[df["商品類別"].notna() & df["商品名稱"].notna()]
        .groupby("商品名稱")["商品類別"]
        .apply(deterministic_mode)
    )
    cat_missing2 = df["商品類別"].isna()
    fillable_cat_by_name = cat_missing2 & df["商品名稱"].map(
        lambda x: x in cat_mode_by_name.index if pd.notna(x) else False
    )
    filled_by_name = 0
    for idx in df.index[fillable_cat_by_name]:
        nm = df.at[idx, "商品名稱"]
        new_cat = cat_mode_by_name.get(nm, np.nan)
        if pd.notna(new_cat):
            change_log.append((df.at[idx, "_原始列號"], "商品類別", "缺值時依商品名稱多數值補齊(次要)",
                                df.at[idx, "商品類別"], new_cat))
            df.at[idx, "商品類別"] = new_cat
            filled_by_name += 1
    log(f"商品類別第二層(依名稱)補齊：成功 {filled_by_name} 筆")

    # ---- 規則5：商品類別缺值且ID與名稱皆無法分類 ----
    rule5_hit = df["商品類別"].isna()
    log(f"規則5 商品類別缺值且無法分類：命中 {int(rule5_hit.sum())} 筆")
    reason[rule5_hit.values & reason.isna()] = "商品類別缺值且無法分類"

    # ---- 規則6：訂購數量缺漏或非正值 ----
    rule6_hit = df["訂購數量_數值"].isna() | (df["訂購數量_數值"] <= 0)
    log(f"規則6 訂購數量缺漏或非正值：命中 {int(rule6_hit.sum())} 筆")
    reason[rule6_hit.values & reason.isna()] = "訂購數量缺漏或非正值"

    # ---- 規則7：單位缺漏或異常 ----
    rule7_hit = df["單位"].isna() | ~df["單位"].map(lambda x: x in UNIT_WHITELIST if pd.notna(x) else False)
    log(f"規則7 單位缺漏或異常：命中 {int(rule7_hit.sum())} 筆")
    reason[rule7_hit.values & reason.isna()] = "單位缺漏或異常"

    # ---- 清理後二次重複列檢查（正規化後14欄完全相同，第一輪原始值比對未偵測到者）----
    normalized_dup_mask = df.duplicated(subset=ORDER_COLUMNS, keep="first")
    newly_caught = normalized_dup_mask.values & reason.isna()
    log(f"清理後二次重複列：命中 {int(newly_caught.sum())} 筆")
    reason[newly_caught] = "清理後二次重複列(NULL字樣與真空白標準化後才顯現的完全重複)"

    # ---- 規則8：商品類別代碼81/82疑似測試資料（人工判讀規則）----
    rule8_hit = df["商品類別"].isin([81, 82]) & reason.isna()
    log(f"規則8 商品類別代碼81/82疑似測試資料：命中 {int(rule8_hit.sum())} 筆")
    reason[rule8_hit] = "疑似測試資料(人工確認刪除)"

    df["_隔離原因"] = reason
    return df, change_log


# ============================================================
# 3. 商品類別對照表擴充比對 + 隔離資料關鍵字重分類
#    （清洗文件.md Part 1 §2-1補、Part 6）
# ============================================================

def resolve_categories_and_reclassify(clean_df: pd.DataFrame, iso_df: pd.DataFrame):
    # 乾淨資料：所有商品類別皆已通過規則5檢驗，理論上不應為空值
    unresolved = clean_df["商品類別"].isna() | ~clean_df["商品類別"].isin(VALID_CATEGORY_IDS)
    if unresolved.any():
        log(f"⚠️ 警告：乾淨資料中仍有 {int(unresolved.sum())} 筆商品類別代碼不在 69 碼對照表內（保留原值，不強制修改）")

    # 隔離資料：針對「商品類別缺值且無法分類」的列，依商品名稱關鍵字規則重分類
    target_mask = iso_df["_隔離原因"] == "商品類別缺值且無法分類"
    log(f"隔離資料中『商品類別缺值且無法分類』共 {int(target_mask.sum())} 筆，開始關鍵字重分類")

    reclass_result = iso_df.loc[target_mask, "商品名稱"].map(classify_by_keyword)
    matched = reclass_result.notna() & (reclass_result != -1)
    test_placeholder = reclass_result == -1

    candidate_idx = reclass_result.index[matched]
    log(f"關鍵字規則命中並可分類：{len(candidate_idx)} 筆")
    log(f"命中『疑似測試資料』佔位字樣（不分類，維持隔離）：{int(test_placeholder.sum())} 筆")
    still_unmatched = int(target_mask.sum()) - len(candidate_idx) - int(test_placeholder.sum())
    log(f"規則未命中，維持隔離待人工確認：{still_unmatched} 筆")

    # ---- Part 11：移回乾淨資料前，重新檢查訂購數量／單位規則 ----
    # 這批列的「隔離主要原因」是依優先序取第一個命中的規則（商品類別缺值，優先序5），
    # 不代表它們沒有同時踩到訂購數量(優先序6)／單位(優先序7)的問題，只是優先序邏輯
    # 沒有把後面命中的規則列為「主要原因」。若只用關鍵字把類別補上就直接移回乾淨資料，
    # 可能讓「訂購數量<=0」或「單位不合法」的列未經檢查混入乾淨資料，違反乾淨資料
    # 「訂購數量100%正整數、單位100%合法」的既有保證，因此類別分類成功後，移動前
    # 一定要重新檢查這兩條規則；仍異常者維持隔離，並把隔離原因改標成真正命中的規則。
    qty = iso_df.loc[candidate_idx, "訂購數量_數值"]
    unit = iso_df.loc[candidate_idx, "單位"]
    qty_bad = qty.isna() | (qty <= 0)
    unit_bad = unit.isna() | ~unit.map(lambda x: x in UNIT_WHITELIST if pd.notna(x) else False)

    recheck_reason = pd.Series(index=candidate_idx, dtype=object)
    recheck_reason[qty_bad] = "訂購數量缺漏或非正值"
    recheck_reason[(~qty_bad) & unit_bad] = "單位缺漏或異常"
    still_bad_idx = recheck_reason.dropna().index
    move_idx = candidate_idx.difference(still_bad_idx)
    if len(still_bad_idx) > 0:
        log(f"關鍵字命中但訂購數量／單位仍異常，維持隔離（改標記真正原因）：{len(still_bad_idx)} 筆")

    iso_df.loc[candidate_idx, "商品類別"] = reclass_result.loc[candidate_idx].astype(int)
    iso_df.loc[move_idx, "_隔離原因"] = "_MOVE_TO_CLEAN_"  # 標記待搬移
    iso_df.loc[still_bad_idx, "_隔離原因"] = recheck_reason.loc[still_bad_idx]

    moved_rows = iso_df.loc[iso_df["_隔離原因"] == "_MOVE_TO_CLEAN_"].copy()
    remaining_iso_df = iso_df.loc[iso_df["_隔離原因"] != "_MOVE_TO_CLEAN_"].copy()

    if len(moved_rows) > 0:
        new_clean_df = pd.concat([clean_df, moved_rows.drop(columns=["_隔離原因"])], ignore_index=True)
    else:
        new_clean_df = clean_df

    # 提供給 /api/clean/summary 的「隔離資料關鍵字重分類」明細統計（對應前端原型 renderCleanResults
    # 的 reclass 區塊）：待重分類列數、成功移回、命中測試佔位字樣、命中但數量/單位仍異常、規則未命中。
    reclass_stats = {
        "target_total": int(target_mask.sum()),
        "moved": int(len(move_idx)),
        "test_placeholder": int(test_placeholder.sum()),
        "still_bad": int(len(still_bad_idx)),
        "still_unmatched": int(still_unmatched),
    }

    return new_clean_df, remaining_iso_df, reclass_stats


# ============================================================
# 3.3 商品名稱品質：非描述性名稱隔離（Part 17）
# ============================================================

def isolate_nondescriptive_names(clean_df: pd.DataFrame, iso_df: pd.DataFrame):
    """把乾淨資料中『商品名稱幾乎沒有可辨識品名文字』（純規格代號，例如海鮮的「(60) 10/15」）
    的列移到隔離資料，隔離原因標為待補正式品名。這種名稱無法從資料內還原（已用原始檔驗證：
    同商品ID每一列都是規格碼），需人工用供應商品名對照表補上正式名稱後再回收。

    放在關鍵字重分類「之後」執行：因為部分規格碼名稱是靠關鍵字（如 ^\\(\\d+\\) → 生鮮）才被救回
    乾淨資料的，若在重分類前就過濾會漏抓；在最終乾淨集合上統一掃一次，不論它是原本就乾淨、
    還是重分類搬回來的，都能一致處理。"""
    mask = clean_df["商品名稱"].map(is_nondescriptive_name)
    n = int(mask.sum())
    log(f"Part 17 商品名稱非描述性（純規格代號，待補正式品名）：命中 {n} 筆"
        f"（{int(clean_df.loc[mask, '商品ID'].nunique())} 個商品ID），移至隔離資料")
    if n == 0:
        return clean_df, iso_df, n
    moved = clean_df[mask].copy()
    moved["_隔離原因"] = "商品名稱疑似非描述性代號(待補正式品名)"
    kept = clean_df[~mask].copy()
    new_iso_df = pd.concat([iso_df, moved], ignore_index=True)
    return kept, new_iso_df, n


# ============================================================
# 3.5 材積清洗（商品-長／商品-寬／商品-高／商品重量，清洗文件.md Part 9）
# ============================================================

def clean_dimensions_and_weight(clean_df: pd.DataFrame, iso_df: pd.DataFrame, change_log):
    """材積四欄清洗：數值化 → 負值取絕對值 → 0視為缺值 → 極端值判定
    → 兩層回填/校正（優先依同商品ID其他正常列眾數；查無則依同商品類別統計範圍嘗試單位校正）
    → 計算材積(cm³)。

    在『乾淨資料＋隔離資料』全體上合併計算：材積是商品本身的物理屬性，跟該列被隔離的
    原因（例如單位欄異常、訂購數量不合法）無關，合併計算可取得最大樣本數，讓「同商品ID」
    「同商品類別」的參考基準更可靠。清洗完成後依 `_原始列號`（全域唯一）拆回兩個
    DataFrame，不依賴列索引對齊（因為 resolve_categories_and_reclassify 已對 clean_df 重編過索引）。
    """
    log("Part 9：材積清洗開始（商品-長／商品-寬／商品-高／商品重量）")
    clean_keys = set(clean_df["_原始列號"])
    iso_keys = set(iso_df["_原始列號"])

    combo = pd.concat([clean_df, iso_df], ignore_index=True, sort=False)

    # ---- 步驟1：數值化 ----
    for col in VOLUME_ALL_COLUMNS:
        combo[f"{col}_原始值"] = combo[col]
        combo[col] = pd.to_numeric(combo[col], errors="coerce")
        parse_fail = combo[f"{col}_原始值"].notna() & combo[col].isna()
        combo[f"{col}_原始無法解析_flag"] = parse_fail.astype(int)
        log(f"  {col}：無法解析為數字 {int(parse_fail.sum())} 筆")

    # ---- 步驟2：負值 → 絕對值（比照訂購數量負值的處理原則）----
    for col in VOLUME_ALL_COLUMNS:
        neg_mask = combo[col] < 0
        n_neg = int(neg_mask.sum())
        if n_neg:
            for idx in combo.index[neg_mask]:
                change_log.append((combo.at[idx, "_原始列號"], col, "材積負值→絕對值",
                                    combo.at[idx, col], abs(combo.at[idx, col])))
            combo.loc[neg_mask, col] = combo.loc[neg_mask, col].abs()
        combo[f"{col}_原為負值_flag"] = neg_mask.astype(int)
        log(f"  {col}：負值 {n_neg} 筆（已轉絕對值）")

    # ---- 步驟3：0 視為缺值（0在物流情境下不具業務意義，比照清洗文件 Part1 §2-2）----
    for col in VOLUME_ALL_COLUMNS:
        zero_mask = combo[col] == 0
        combo[f"{col}_原始為0_flag"] = zero_mask.astype(int)
        combo.loc[zero_mask, col] = np.nan
        log(f"  {col}：原始為0（轉缺值） {int(zero_mask.sum())} 筆")
    combo["尺寸重量_原始為0_flag"] = combo[[f"{c}_原始為0_flag" for c in VOLUME_ALL_COLUMNS]].max(axis=1)

    # ---- 步驟4：極端值判定（校正前的初始判斷）----
    for col in VOLUME_ALL_COLUMNS:
        combo[f"{col}_極端值_flag"] = (combo[col] > EXTREME_MAX[col]).astype(int)
        log(f"  {col}：極端值(>{EXTREME_MAX[col]}) {int(combo[f'{col}_極端值_flag'].sum())} 筆")

    # ---- 步驟5：建立參考基準（僅用「有效樣本」＝非缺值且非極端值的列）----
    id_mode_ref, cat_stats_ref = {}, {}
    for col in VOLUME_ALL_COLUMNS:
        valid_mask = combo[col].notna() & (combo[f"{col}_極端值_flag"] == 0)
        combo[f"{col}__valid"] = valid_mask
        valid = combo.loc[valid_mask, ["商品ID", "商品類別", col]]
        id_mode_ref[col] = valid.groupby("商品ID")[col].apply(deterministic_mode)
        cat_group = valid.groupby("商品類別")[col]
        cat_stats_ref[col] = pd.DataFrame({"median": cat_group.median(), "n": cat_group.count()})

    # ---- 步驟6：兩層回填/校正 ----
    for col in VOLUME_ALL_COLUMNS:
        valid_mask = combo[f"{col}__valid"]
        was_extreme = combo[f"{col}_極端值_flag"] == 1
        need_fix = ~valid_mask & (combo[col].isna() | was_extreme)
        original_for_log = combo[col].copy()

        # 第一層：同商品ID其他正常列的眾數（等同於用「同一個商品」的其他紀錄回填，
        # 因為此資料集商品ID與商品名稱為一對一，用ID分組即等於用名稱分組）
        id_fill_val = combo["商品ID"].map(id_mode_ref[col])
        use_id_fill = need_fix & id_fill_val.notna()

        # 第1.5層（Part 10 新增）：共用可疑值偵測——同一個已判定極端值的原始數值，
        # 若重複出現在異常多個不同商品ID上，視為系統共用預設值/測試值，不採信、不嘗試校正，
        # 直接當缺值處理（避免被下面的類別校正法誤判為「已修正的合理值」）
        extreme_val_by_sku = (
            combo.loc[was_extreme & combo[col].notna()].drop_duplicates("商品ID")[col]
        )
        shared_val_counts = extreme_val_by_sku.value_counts()
        shared_suspect_vals = set(shared_val_counts[shared_val_counts >= SHARED_VALUE_MIN_DISTINCT_SKU].index)
        is_shared_suspect = was_extreme & combo[col].isin(shared_suspect_vals)
        if shared_suspect_vals:
            n_sku_affected = int(extreme_val_by_sku.isin(shared_suspect_vals).sum())
            log(f"  {col}：偵測到共用可疑極端值 {sorted(shared_suspect_vals)}，"
                f"分布於 {n_sku_affected} 個不同商品ID，命中 {int(is_shared_suspect.sum())} 筆，視為缺值不予校正")

        # 第二層：只對「原本是極端值」（非單純缺值/0）、非共用可疑值、且第一層查無結果者，
        # 嘗試依同商品類別統計範圍做單位校正（÷10／÷100／÷1000），優先採用除數最小(校正幅度最小)的候選
        cat_median = combo["商品類別"].map(cat_stats_ref[col]["median"])
        cat_n = combo["商品類別"].map(cat_stats_ref[col]["n"])
        try_cat_fix = (need_fix & ~id_fill_val.notna() & was_extreme & ~is_shared_suspect
                       & cat_n.fillna(0).ge(CATEGORY_MIN_SAMPLE) & cat_median.notna())

        best_value = pd.Series(np.nan, index=combo.index)
        best_factor = pd.Series(np.nan, index=combo.index)
        lo = cat_median * CATEGORY_PLAUSIBLE_RATIO[0]
        hi = cat_median * CATEGORY_PLAUSIBLE_RATIO[1]
        # Part 13：類別校正候選除了要落在「同類別中位數的1/3～3倍」區間內，
        # 還必須「本身不再超過該欄的極端值門檻」（EXTREME_MAX）。否則對中位數本來就偏大的
        # 類別（例如中位數100cm，3倍上限=300cm），一個÷10候選可能落在200~300cm之間仍被
        # 視為「已成功校正」，把校正來源標成正常、極端值旗標清為0、材積待複核＝0，等於讓一個
        # 仍>門檻的可疑值靜默混入乾淨資料（重量同理：中位數20kg×3=60kg，可能放行>30kg的校正值）。
        # 加上這道上限，任何校正結果一律不超過門檻；被卡住的極端值會維持「無法處理」→材積待複核＝1。
        for factor in UNIT_CORRECTION_FACTORS:
            candidate = combo[col] / factor
            fits = (try_cat_fix & candidate.between(lo, hi)
                    & candidate.le(EXTREME_MAX[col]) & best_value.isna())
            best_value[fits] = candidate[fits]
            best_factor[fits] = factor
        use_cat_fix = try_cat_fix & best_value.notna()

        for idx in combo.index[use_id_fill]:
            change_log.append((combo.at[idx, "_原始列號"], col,
                                "材積缺值/0值/極端值→依同商品ID其他正常列眾數回填",
                                original_for_log.at[idx], id_fill_val.at[idx]))
        combo.loc[use_id_fill, col] = id_fill_val[use_id_fill]
        combo.loc[use_id_fill, f"{col}_校正來源"] = "同商品ID眾數"
        combo.loc[use_id_fill, f"{col}_極端值_flag"] = 0

        for idx in combo.index[use_cat_fix]:
            factor_int = int(best_factor.at[idx])
            change_log.append((combo.at[idx, "_原始列號"], col,
                                f"材積極端值→依商品類別統計範圍校正(÷{factor_int})",
                                original_for_log.at[idx], best_value.at[idx]))
        combo.loc[use_cat_fix, col] = best_value[use_cat_fix]
        combo.loc[use_cat_fix, f"{col}_校正來源"] = best_factor[use_cat_fix].map(lambda f: f"依類別範圍校正(÷{int(f)})")
        combo.loc[use_cat_fix, f"{col}_極端值_flag"] = 0

        combo.loc[valid_mask, f"{col}_校正來源"] = "原始正常值"

        # 共用可疑值：未被同商品ID眾數救回者（Part 10 新增），直接視為缺值，不編造校正數字
        shared_unresolved = is_shared_suspect & ~use_id_fill
        for idx in combo.index[shared_unresolved]:
            change_log.append((combo.at[idx, "_原始列號"], col,
                                "材積極端值→疑似多商品共用預設值，改為缺值",
                                original_for_log.at[idx], np.nan))
        combo.loc[shared_unresolved, col] = np.nan
        combo.loc[shared_unresolved, f"{col}_校正來源"] = "無法處理(疑似共用預設值,已排除)"

        unresolved_na = need_fix & ~use_id_fill & ~use_cat_fix & ~shared_unresolved & combo[col].isna()
        unresolved_extreme = need_fix & ~use_id_fill & ~use_cat_fix & ~shared_unresolved & combo[col].notna()
        combo.loc[unresolved_na, f"{col}_校正來源"] = "無法處理(維持缺值)"
        combo.loc[unresolved_extreme, f"{col}_校正來源"] = "無法處理(維持原始極端值,待人工複核)"

        combo[f"{col}_共用可疑值_flag"] = is_shared_suspect.astype(int)

        log(f"  {col}：依商品ID眾數回填 {int(use_id_fill.sum())} 筆／依類別範圍校正 {int(use_cat_fix.sum())} 筆／"
            f"共用可疑值改缺值 {int(shared_unresolved.sum())} 筆／無法處理 {int((unresolved_na | unresolved_extreme).sum())} 筆")

    # ---- 步驟6.5（Part 15 新增）：長寬高缺值 → 同商品類別中位數推估（僅長/寬/高，不含重量）----
    # 尺寸是商品屬性、同商品ID每筆訂單都一樣，所以同ID/同名互補救不了缺值（步驟6實測回填0筆）。
    # 唯一能自動補的是「同類別、其他有實測值商品」的中位數。為避免高訂單量商品灌爆中位數，
    # 類別中位數以「每個商品一個代表值」計算（先收斂到商品層級再取中位數）。此時欄位裡所有
    # notna 的值都還是實測/校正、尚無推估值，因此可安全拿來當推估基準。每個補上的值標記
    # 校正來源="類別中位數推估"，不與實測混淆；仍無類別參考者維持缺值，交由人工（見報告建議）。
    for col in DIM_COLUMNS:
        still_missing = combo[col].isna()
        if not bool(still_missing.any()):
            continue
        measured = combo.loc[combo[col].notna(), ["商品ID", "商品類別", col]]
        # 商品層級代表值：尺寸在同商品內為常數，取中位數即等於該商品的值
        prod = measured.groupby("商品ID").agg(val=(col, "median"), cat=("商品類別", "first"))
        prod = prod.dropna(subset=["val", "cat"])
        cat_grp = prod.groupby("cat")["val"]
        cat_median = cat_grp.median()
        cat_prod_n = cat_grp.count()  # 該類別「有實測值的不同商品數」
        ok_cats = set(cat_prod_n[cat_prod_n >= IMPUTE_MIN_CATEGORY_SAMPLE].index)

        med_map = combo["商品類別"].map(cat_median)
        n_map = combo["商品類別"].map(cat_prod_n).fillna(0)
        do_impute = still_missing & med_map.notna() & combo["商品類別"].isin(ok_cats)
        combo.loc[do_impute, col] = med_map[do_impute]
        combo.loc[do_impute, f"{col}_校正來源"] = "類別中位數推估"

        n_rows = int(do_impute.sum())
        n_prod = int(combo.loc[do_impute, "商品ID"].nunique())
        still_blank = still_missing & ~do_impute
        n_blank_prod = int(combo.loc[still_blank, "商品ID"].nunique())
        log(f"  {col}：類別中位數推估補值 {n_rows} 列（{n_prod} 個商品）；"
            f"仍無類別參考維持缺值 {int(still_blank.sum())} 列（{n_blank_prod} 個商品）")

    # ---- 步驟7：材積(cm³)＝長×寬×高（三邊皆有值才計算）----
    combo["材積_cm3"] = np.where(
        combo[DIM_COLUMNS].notna().all(axis=1),
        combo["商品-長"] * combo["商品-寬"] * combo["商品-高"],
        np.nan,
    )

    # 材積本身的極端值判定（Part 10 新增）：長寬高各自都可能<200cm門檻個別過關，
    # 但三者相乘後的體積仍可能不合理（實測案例：102×122×160cm，各自未超標但體積達199萬cm³，
    # 且經測試同商品ID/同類別校正法皆無法安全救回，見清洗文件附註）
    combo["材積_cm3_極端值_flag"] = (
        combo["材積_cm3"].notna() & (combo["材積_cm3"] > VOLUME_EXTREME_MAX)
    ).astype(int)
    n_vol_extreme = int(combo["材積_cm3_極端值_flag"].sum())
    log(f"  材積_cm3：體積極端值(>{VOLUME_EXTREME_MAX:,.0f}) {n_vol_extreme} 筆"
        f"（單一維度未超標但體積不合理，無法安全校正，將於商品材積主檔標記材積不可用）")

    combo["尺寸重量_極端值_flag"] = combo[
        [f"{c}_極端值_flag" for c in VOLUME_ALL_COLUMNS] + ["材積_cm3_極端值_flag"]
    ].max(axis=1)

    # ---- 建立商品材積主檔（一個商品ID一列，供後續貨架／棧板判斷使用）----
    sku_dim_df = build_sku_dimension_master(combo)

    # 把商品層級的儲位分類（貨架/棧板）對映回每一列，讓下游可直接在乾淨/隔離資料篩選（Part 16）
    _cls_map = sku_dim_df.set_index("商品ID")["儲位分類"]
    combo["儲位分類"] = combo["商品ID"].map(_cls_map).fillna("未分類(無商品ID)")

    # ---- Part 11：把24欄逐維度flag/校正來源，濃縮成單一「材積待複核」欄 ----
    # 原本每個維度各自輸出6種欄位，對104萬列的主表而言，「已成功回填或校正」「原始正常值」
    # 的列這些欄位99%以上都是0，逐列攤開反而讓人看不出真正需要注意的是哪幾筆；真正該被看見
    # 的，是唯一「有問題卻沒被自動修好、原始可疑數值仍原樣留在欄位裡」的狀態（校正來源＝
    # 「無法處理(維持原始極端值,待人工複核)」），因為這種列在精簡後的資料表裡長得跟正常資料
    # 一模一樣，若不標記，後續拿材積做「貨架區 vs 棧板區」判斷時會被誤當成真實尺寸。因此改為
    # 單一彙總flag：任一維度「無法處理，維持原始極端值」，或長寬高相乘後的體積本身超標
    # （見 VOLUME_EXTREME_MAX），就標記為1；已成功回填/校正的列、原始就正常的列都是0。
    # 完整逐維度細節（原始值、校正方式、除數）不會因此遺失，仍完整保留在「逐筆變更紀錄」
    # （成功校正/回填的逐筆紀錄）與「商品材積主檔」的資料來源分布欄（彙總視角）裡。
    # 材積來源（Part 15 新增）：三維皆有值時，任一維來自類別中位數推估 →「類別推估」，
    # 否則「實測」；任一維仍缺（算不出體積）→「無法計算」。供下游貨架/棧板判斷分辨可信度。
    any_imputed = pd.Series(False, index=combo.index)
    for col in DIM_COLUMNS:
        any_imputed |= (combo[f"{col}_校正來源"] == "類別中位數推估")
    dims_all_present = combo[DIM_COLUMNS].notna().all(axis=1)
    combo["材積來源"] = np.where(
        ~dims_all_present, "無法計算",
        np.where(any_imputed, "類別推估", "實測"),
    )
    src_counts = pd.Series(combo["材積來源"]).value_counts().to_dict()
    log(f"材積來源分布：{src_counts}")

    # 材積待複核：僅看長/寬/高（本專案不使用商品重量，故重量極端值不納入此旗標），
    # 條件＝任一維度「無法處理，維持原始極端值」，或長寬高相乘後的體積本身超標。
    unresolved_extreme_any = pd.Series(False, index=combo.index)
    for col in DIM_COLUMNS:
        unresolved_extreme_any |= (combo[f"{col}_校正來源"] == "無法處理(維持原始極端值,待人工複核)")
    combo["材積待複核"] = (unresolved_extreme_any | (combo["材積_cm3_極端值_flag"] == 1)).astype(int)
    log(f"材積待複核（長寬高無法自動校正、原始可疑極端值仍原樣保留，或體積超標）：{int(combo['材積待複核'].sum())} 筆，"
        f"已在乾淨/隔離資料以單一「材積待複核」欄位標示，逐維度細節見商品材積主檔／逐筆變更紀錄")

    # ---- 清除暫存欄位，依 _原始列號 拆回原本兩個 DataFrame ----
    # 注意：iso_df 獨有的 "_隔離原因" 欄位在 concat 時會被自動補上（乾淨資料那側全為NaN），
    # 拆回 new_clean_df 時必須把它丟掉，否則乾淨資料會多出一欄全空值的「_隔離原因」。
    # 同時把24欄逐維度flag/校正來源也一併丟掉（已濃縮進上面的「材積待複核」），主表只保留
    # 「材積_cm3」與「材積待複核」兩欄，避免大量恆為0的稀疏欄位淹沒真正需要注意的列。
    drop_cols = (
        [f"{c}_原始值" for c in VOLUME_ALL_COLUMNS]
        + [f"{c}__valid" for c in VOLUME_ALL_COLUMNS]
        + MATERIAL_DETAIL_COLUMNS
    )
    combo = combo.drop(columns=drop_cols)
    new_clean_df = (
        combo[combo["_原始列號"].isin(clean_keys)]
        .drop(columns=["_隔離原因"], errors="ignore")
        .reset_index(drop=True)
    )
    new_iso_df = combo[combo["_原始列號"].isin(iso_keys)].reset_index(drop=True)
    log(f"Part 9：材積清洗完成，商品材積主檔共 {len(sku_dim_df)} 個商品ID")

    return new_clean_df, new_iso_df, sku_dim_df, change_log


def classify_storage(vol_cm3, max_edge_cm, size_source):
    """依商品尺寸判定儲位分類（Part 16）：回傳 (儲位分類, 分類依據)。
    缺尺寸者暫歸棧板待量測；材積>門檻或最長邊>門檻者歸棧板；其餘歸貨架。"""
    if size_source == "無法計算(缺維度)":
        return "棧板", "缺尺寸暫歸棧板(待量測)"
    reasons = []
    if pd.notna(vol_cm3) and vol_cm3 > SHELF_MAX_VOLUME_CM3:
        reasons.append(f"材積{vol_cm3/1000:.0f}L>{SHELF_MAX_VOLUME_CM3/1000:.0f}L")
    if pd.notna(max_edge_cm) and max_edge_cm > SHELF_MAX_EDGE_CM:
        reasons.append(f"最長邊{max_edge_cm:.0f}cm>{SHELF_MAX_EDGE_CM:.0f}cm")
    if reasons:
        return "棧板", ";".join(reasons)
    return "貨架", f"材積≤{SHELF_MAX_VOLUME_CM3/1000:.0f}L且最長邊≤{SHELF_MAX_EDGE_CM:.0f}cm"


def build_sku_dimension_master(combo: pd.DataFrame) -> pd.DataFrame:
    """一個商品ID一列的材積主檔：取該ID全部列（清洗/校正後）的眾數作代表值，
    並標註資料來源分布、多重數值衝突、與貨架/棧板儲位分類，供後續 ABC 分組與儲位配置使用。

    效能：原本用 `for pid, g in combo.groupby("商品ID")` 逐商品跑一輪，商品內又對名稱/
    類別/長/寬/高/重量各呼叫一次 deterministic_mode()、材積來源欄各呼叫一次 value_counts()、
    長寬高重量各呼叫一次 nunique()——商品數（例：5,672）× 十幾個操作，每次都要從百萬列全表
    重新切出該商品的子表，pandas 逐次呼叫的固定開銷疊加起來，實測是材積清洗最慢的一段（在
    正式資料集上，這段加上前面的欄位計算共佔材積清洗約七成時間）。改成用 groupby 的原生
    聚合（.apply／.nunique／.size／.any 等）一次算完所有商品，數量級相同的結果、快非常多；
    只有 classify_storage() 因為要組裝可讀文字、且只在「一個商品一列」的小表（商品數等級，
    不是百萬列）上跑，維持逐列呼叫，不構成效能問題。

    唯一不完全等價之處：「長/寬/高/重量_資料來源分布」欄位在多種來源筆數剛好相同（tie）時
    的顯示順序，原本依賴 pandas value_counts() 對 tie 的內部雜湊順序（本身就不是有文件保證
    的行為），這裡改用「筆數由多到少，筆數相同者依名稱排序」的確定性規則。純顯示用的稽核
    文字欄位，數字本身（每種來源各幾筆）完全不變，不影響任何下游判斷（儲位分類／材積可用／
    代表值等欄位皆已驗證與原版逐商品版本逐列完全一致）。"""
    g = combo.groupby("商品ID")

    name = g["商品名稱"].apply(deterministic_mode)
    cat = g["商品類別"].apply(deterministic_mode)
    length = g["商品-長"].apply(deterministic_mode)
    width = g["商品-寬"].apply(deterministic_mode)
    height = g["商品-高"].apply(deterministic_mode)
    weight = g[WEIGHT_COLUMN].apply(deterministic_mode)
    idx = length.index  # 統一以此為準對齊各聚合結果（皆來自同一個 groupby("商品ID")）

    dims_present = length.notna() & width.notna() & height.notna()
    vol = pd.Series(np.where(dims_present, length * width * height, np.nan), index=idx)

    # 各欄「校正來源」分布：用 groupby(["商品ID", 來源欄]) 一次算出全部商品×來源的筆數，
    # 取代逐商品呼叫 value_counts()；再依商品把小小的 (通常1~3種來源) 子結果組成顯示字串。
    def _fmt_source_counts(s: pd.Series) -> str:
        # s：某商品的「來源→筆數」小 Series（拿掉外層商品ID索引後只剩來源名稱）。
        # 先依來源名稱排序，再用穩定排序依筆數由多到少排——筆數相同者因為穩定排序
        # 會保留前一步的字母序，等於「筆數多到少，同筆數依名稱排序」，結果每次都一樣。
        s = s.droplevel(0).sort_index()
        s = s.sort_values(ascending=False, kind="mergesort")
        return "; ".join(f"{k}×{v}" for k, v in s.items())

    來源分布 = {}
    for col in VOLUME_ALL_COLUMNS:
        vc = combo.groupby(["商品ID", f"{col}_校正來源"]).size()
        來源分布[col] = vc.groupby(level=0).apply(_fmt_source_counts).reindex(idx)

    conflict = {
        col: (g[col].nunique() > 1).astype(int).reindex(idx)
        for col in ("商品-長", "商品-寬", "商品-高", WEIGHT_COLUMN)
    }

    # 尺寸來源（Part 15）：只看長/寬/高——三維皆有值時，任一維任一列出現「類別中位數推估」
    # → 類別推估；否則 實測；任一維仍缺（算不出材積）→ 無法計算(缺維度)。
    imputed_row = combo[[f"{c}_校正來源" for c in DIM_COLUMNS]].eq("類別中位數推估").any(axis=1)
    imputed = imputed_row.groupby(combo["商品ID"]).any().reindex(idx, fill_value=False)
    尺寸來源 = pd.Series(
        np.where(~dims_present, "無法計算(缺維度)", np.where(imputed, "類別推估", "實測")),
        index=idx,
    )

    max_edge = pd.concat([length, width, height], axis=1).max(axis=1, skipna=True)
    資料列數 = g.size().reindex(idx)

    sku_df = pd.DataFrame({
        "商品ID": idx,
        "商品名稱": name.reindex(idx).values,
        "商品類別": cat.reindex(idx).values,
        "代表長cm": length.values,
        "代表寬cm": width.reindex(idx).values,
        "代表高cm": height.reindex(idx).values,
        "代表重量kg": weight.reindex(idx).values,
        "代表材積cm3": vol.values,
        "最長邊cm": max_edge.values,
        "尺寸來源": 尺寸來源.values,
        "材積可用": vol.notna().values,
        "資料列數": 資料列數.values,
        "長_資料來源分布": 來源分布["商品-長"].values,
        "寬_資料來源分布": 來源分布["商品-寬"].values,
        "高_資料來源分布": 來源分布["商品-高"].values,
        "重量_資料來源分布": 來源分布[WEIGHT_COLUMN].values,
        "長_同ID多重數值衝突": conflict["商品-長"].values,
        "寬_同ID多重數值衝突": conflict["商品-寬"].values,
        "高_同ID多重數值衝突": conflict["商品-高"].values,
        "重量_同ID多重數值衝突": conflict[WEIGHT_COLUMN].values,
    })

    # 儲位分類（Part 16）：僅在「一個商品一列」的小表上逐列呼叫 classify_storage()，
    # 商品數量級（例：5,672），不是百萬列，不構成效能問題。
    if len(sku_df):
        儲位分類, 分類依據 = zip(*(
            classify_storage(v, m, s)
            for v, m, s in zip(sku_df["代表材積cm3"], sku_df["最長邊cm"], sku_df["尺寸來源"])
        ))
    else:
        儲位分類, 分類依據 = (), ()
    sku_df["儲位分類"] = list(儲位分類)
    sku_df["分類依據"] = list(分類依據)

    return sku_df[[
        "商品ID", "商品名稱", "商品類別", "代表長cm", "代表寬cm", "代表高cm", "代表重量kg",
        "代表材積cm3", "最長邊cm", "尺寸來源", "儲位分類", "分類依據", "材積可用", "資料列數",
        "長_資料來源分布", "寬_資料來源分布", "高_資料來源分布", "重量_資料來源分布",
        "長_同ID多重數值衝突", "寬_同ID多重數值衝突", "高_同ID多重數值衝突", "重量_同ID多重數值衝突",
    ]].reset_index(drop=True)


# ============================================================
# 3.6 配送地址補齊（縣市／區域，Part 18，供區域併單策略使用）
# ============================================================

def backfill_delivery_address(clean_df: pd.DataFrame, iso_df: pd.DataFrame, change_log):
    """配送地址補齊縣市/區域，供區域併單策略使用。規則細節見檔案開頭 Part 18 更新記錄。

    與材積清洗（Part 9）相同，在『乾淨資料＋隔離資料』全體上合併計算：配送地址是門市的
    屬性，跟該列被隔離的原因無關，合併計算才能取得每個門市代碼最大的可信地址樣本數。
    清洗完成後依 `_原始列號`（全域唯一）拆回兩個 DataFrame。
    """
    log("Part 18：配送地址補齊開始（縣市／區域，供區域併單使用）")
    clean_keys = set(clean_df["_原始列號"])
    iso_keys = set(iso_df["_原始列號"])

    combo = pd.concat([clean_df, iso_df], ignore_index=True, sort=False)

    # ---- 步驟1：地址前綴回推＋新舊制行政區更名正規化（無論該列地址是否完整，一律套用）。
    #     Part 22 起，normalize_address_prefix() 除了原有的新舊制縣市更名，還會先去除
    #     誤帶的郵遞區號、展開常見城市簡寫、並對省略縣名但寫出全國唯一鄉鎮市名稱者
    #     （如屏東市、琉球鄉）回推補上所屬縣，詳見該函式與 TOWNSHIP_TO_COUNTY 內部註解 ----
    combo["_addr_norm"] = combo[ADDRESS_COLUMN].map(normalize_address_prefix)
    norm_changed = combo[ADDRESS_COLUMN].notna() & (combo[ADDRESS_COLUMN] != combo["_addr_norm"])
    log(f"  地址前綴回推＋新舊制行政區更名正規化：命中 {int(norm_changed.sum())} 筆")

    # ---- 步驟2：出貨日期（僅供本函式內部日期比對使用；最終輸出的「出貨日期」欄位
    #     仍由 add_shipdate_column 於輸出階段統一產生，兩者計算方式相同，結果一致）----
    combo["_ship_date_tmp"] = pd.to_datetime(combo["系統出貨單號"].map(parse_shipno_date))

    # ---- 步驟3：拆解縣市／區域／其餘路名門牌 ----
    split_result = combo["_addr_norm"].map(split_city_district_rest)
    combo["_city"] = split_result.map(lambda t: t[0])
    combo["_district"] = split_result.map(lambda t: t[1])
    combo["_rest"] = split_result.map(lambda t: t[2])
    combo["_is_blank"] = combo[ADDRESS_COLUMN].isna() | (combo["_addr_norm"].fillna("") == "")
    combo["_is_full"] = combo["_city"].notna() & combo["_district"].notna() & ~combo["_is_blank"]

    # ---- 步驟4：共用預設值偵測（比照 Part 10 材積清洗 SHARED_VALUE_MIN_DISTINCT_SKU 的手法，
    #     但只排除「稀疏配對」，理由見上方常數區塊的實測發現說明）----
    full_rows = combo.loc[combo["_is_full"]]
    pair_counts = (
        full_rows.groupby(["_addr_norm", STORE_CODE_COLUMN]).size().reset_index(name="_cnt")
    )
    sparse_pairs = pair_counts[pair_counts["_cnt"] <= SHARED_ADDR_SPARSE_MAX_ROWS_PER_CODE]
    sparse_code_count_per_addr = sparse_pairs.groupby("_addr_norm")[STORE_CODE_COLUMN].nunique()
    suspect_addrs = set(sparse_code_count_per_addr[sparse_code_count_per_addr >= SHARED_ADDR_MIN_DISTINCT_CODES].index)
    suspect_sparse_pairs = sparse_pairs.loc[sparse_pairs["_addr_norm"].isin(suspect_addrs)]
    suspect_pair_keys = set(zip(suspect_sparse_pairs["_addr_norm"], suspect_sparse_pairs[STORE_CODE_COLUMN]))
    combo["_addr_code_key"] = list(zip(combo["_addr_norm"], combo[STORE_CODE_COLUMN]))
    combo["_is_placeholder"] = combo["_addr_code_key"].isin(suspect_pair_keys)
    if suspect_pair_keys:
        n_codes_affected = len({code for _, code in suspect_pair_keys})
        n_rows_affected = int(combo["_is_placeholder"].sum())
        # 效能：suspect_addrs 理論上可能有很多筆（例如資料中大量門市共用少數地址時），
        # 全部 sorted() 後塞進一行 log 字串會讓這行印出超長文字，實測在筆數多時本身
        # 就會拖慢清洗（字串組裝＋主控台/日誌輸出 I/O）。改成只列前 20 筆＋總數，
        # 不影響判斷邏輯（suspect_addrs／suspect_pair_keys 的實際使用不變），只是不把整份
        # 清單塞進日誌。
        addr_preview = sorted(suspect_addrs)
        preview_str = "、".join(addr_preview[:20])
        if len(addr_preview) > 20:
            preview_str += f" …等共 {len(addr_preview)} 個地址"
        log(f"  共用預設值偵測：{preview_str}，"
            f"分布於 {n_codes_affected} 個門市代碼的稀疏配對(每代碼<={SHARED_ADDR_SPARSE_MAX_ROWS_PER_CODE}筆)，"
            f"命中 {n_rows_affected} 筆，這些配對不採信為可信地址來源(該地址下其餘高筆數代碼不受影響)")
    # 可信地址：地址完整、非(該代碼稀疏使用的)共用預設值配對、門市代碼本身也不缺值
    combo["_is_trusted"] = combo["_is_full"] & ~combo["_is_placeholder"] & combo[STORE_CODE_COLUMN].notna()

    # ---- 步驟5：建立每個門市代碼的「可信版本」清單（依縣市+區域+其餘路名門牌分組，
    #     含各版本出貨日期區間），依可信版本數分成「單一版本」（向量化處理，占多數）
    #     與「多版本」（疑似搬遷，逐列依出貨日期比對，筆數少很多）兩種情況 ----
    trusted = combo.loc[combo["_is_trusted"], [STORE_CODE_COLUMN, "_city", "_district", "_rest", "_ship_date_tmp"]]
    variant_stats = (
        trusted.groupby([STORE_CODE_COLUMN, "_city", "_district", "_rest"])["_ship_date_tmp"]
        .agg(["min", "max", "count"]).reset_index()
    )
    n_variants_per_code = variant_stats.groupby(STORE_CODE_COLUMN).size()
    single_variant_codes = set(n_variants_per_code[n_variants_per_code == 1].index)
    multi_variant_codes = set(n_variants_per_code[n_variants_per_code > 1].index)

    single_variant_lookup = {}
    multi_variant_lookup = defaultdict(list)
    for r in variant_stats.to_dict("records"):
        code = r[STORE_CODE_COLUMN]
        entry = {"city": r["_city"], "district": r["_district"], "rest": r["_rest"],
                  "min": r["min"], "max": r["max"], "count": r["count"]}
        if code in single_variant_codes:
            single_variant_lookup[code] = entry
        else:
            multi_variant_lookup[code].append(entry)
    log(f"  可信地址門市代碼共 {len(n_variants_per_code)} 個："
        f"單一版本 {len(single_variant_codes)} 個／多版本(疑似搬遷或資料矛盾) {len(multi_variant_codes)} 個")

    need_fill = ~combo["_is_trusted"]
    combo["配送地址_補值來源"] = ""
    combo["配送地址_待確認_flag"] = 0
    original_addr = combo[ADDRESS_COLUMN].copy()

    # ---- 已是可信完整地址：套用步驟1的正規化文字，路名門牌本身完全不動 ----
    combo.loc[combo["_is_trusted"], ADDRESS_COLUMN] = combo.loc[combo["_is_trusted"], "_addr_norm"]
    combo.loc[combo["_is_trusted"] & norm_changed, "配送地址_補值來源"] = "地址前綴回推/新舊制行政區更名正規化(路名門牌未更動)"
    combo.loc[combo["_is_trusted"] & ~norm_changed, "配送地址_補值來源"] = "原始完整(未變更)"

    fill_single_mask = need_fill & combo[STORE_CODE_COLUMN].isin(single_variant_codes)
    fill_multi_mask = need_fill & combo[STORE_CODE_COLUMN].isin(multi_variant_codes)
    fill_unresolved_mask = need_fill & ~fill_single_mask & ~fill_multi_mask

    # ---- 情況A：門市代碼僅有單一可信版本（占需補值列的多數，向量化處理）----
    idx_single = combo.index[fill_single_mask]
    if len(idx_single) > 0:
        sub = combo.loc[idx_single]
        city_dict = {c: v["city"] for c, v in single_variant_lookup.items()}
        district_dict = {c: v["district"] for c, v in single_variant_lookup.items()}
        rest_dict = {c: v["rest"] for c, v in single_variant_lookup.items()}
        v_city = sub[STORE_CODE_COLUMN].map(city_dict)
        v_district = sub[STORE_CODE_COLUMN].map(district_dict)
        v_rest = sub[STORE_CODE_COLUMN].map(rest_dict)

        own_rest_blank = sub["_rest"].isna() | (sub["_rest"].astype(str).str.strip() == "")
        use_full_variant = sub["_is_placeholder"] | sub["_is_blank"] | own_rest_blank

        filled_full = v_city + v_district + v_rest
        filled_partial = v_city + sub["_district"].fillna(v_district) + sub["_rest"].fillna("")
        combo.loc[idx_single, ADDRESS_COLUMN] = filled_partial.where(~use_full_variant, filled_full)
        combo.loc[idx_single, "配送地址_補值來源"] = "同門市代碼單一可信版本回填"

    # ---- 情況B：門市代碼有多個可信版本（實測多為門市搬遷），依該列自己的出貨日期，
    #     採用日期區間最接近的版本（筆數遠少於情況A，逐列處理可接受）----
    idx_multi = combo.index[fill_multi_mask]
    for idx in idx_multi:
        code = combo.at[idx, STORE_CODE_COLUMN]
        variants = multi_variant_lookup[code]
        row_date = combo.at[idx, "_ship_date_tmp"]
        if pd.isna(row_date):
            chosen = max(variants, key=lambda v: v["count"])
        else:
            def _distance(v, _row_date=row_date):
                if pd.notna(v["min"]) and pd.notna(v["max"]) and v["min"] <= _row_date <= v["max"]:
                    return pd.Timedelta(0)
                dists = [abs(_row_date - d) for d in (v["min"], v["max"]) if pd.notna(d)]
                return min(dists) if dists else pd.Timedelta.max
            chosen = min(variants, key=_distance)

        own_rest = combo.at[idx, "_rest"]
        use_full_variant = (combo.at[idx, "_is_placeholder"] or combo.at[idx, "_is_blank"]
                             or not isinstance(own_rest, str) or own_rest.strip() == "")
        if use_full_variant:
            new_addr = f"{chosen['city']}{chosen['district']}{chosen['rest']}"
        else:
            new_district = combo.at[idx, "_district"] if pd.notna(combo.at[idx, "_district"]) else chosen["district"]
            new_addr = f"{chosen['city']}{new_district}{own_rest}"
        combo.at[idx, ADDRESS_COLUMN] = new_addr
        combo.at[idx, "配送地址_補值來源"] = (
            f"同門市代碼多可信版本(疑似搬遷)依出貨日期最接近回填(採用{chosen['min']}~{chosen['max']}期間版本)"
        )

    # ---- 情況C：門市代碼查無可信版本（含代碼本身缺值、地址固定為非地址文字如「甲配」、
    #     或所有列皆命中共用預設值）：保留原值（僅套用步驟1正規化）不變，標記待確認 ----
    idx_unresolved = combo.index[fill_unresolved_mask]
    combo.loc[idx_unresolved, ADDRESS_COLUMN] = combo.loc[idx_unresolved, "_addr_norm"]
    combo.loc[idx_unresolved, "配送地址_補值來源"] = "無法處理(門市代碼查無可信地址,待人工確認)"
    combo.loc[idx_unresolved, "配送地址_待確認_flag"] = 1

    log(f"  同代碼單一可信版本回填：{len(idx_single)} 筆／"
        f"依出貨日期挑選多版本回填：{len(idx_multi)} 筆／"
        f"查無可信版本(待確認)：{len(idx_unresolved)} 筆")

    # ---- 新增「配送區域」欄（Part 21）：只含縣市＋區域（例：新北市內湖區），不含路名
    #     門牌細節，供區域併單策略直接依此欄分組，不必自行從完整地址字串解析。依上面
    #     已完成補值/正規化的最終地址重新拆解，與 ADDRESS_COLUMN 最終結果保持一致 ----
    final_split = combo[ADDRESS_COLUMN].map(split_city_district_rest)
    combo["配送區域"] = final_split.map(
        lambda t: (t[0] + t[1]) if pd.notna(t[0]) and pd.notna(t[1]) else np.nan
    )
    # Part 22 複核修正：情況C（查無可信版本）保留的地址文字，即使表面上縣市/區域
    # 字串完整可解析（例如命中共用預設值偵測的「高雄市橋頭區新莊村大學南路1號」——
    # 文字本身完整，但系統判斷這是跨多個門市代碼共用的可疑預設值，不是該門市代碼
    # 真正的地址），也不代表這是該門市代碼可信的真實地址，此欄一律強制留空，
    # 與 配送地址_待確認_flag=1 保持一致，避免『待確認』卻仍顯示看似可用的區域、
    # 誤被下游併單邏輯直接採用 ----
    combo.loc[combo["配送地址_待確認_flag"] == 1, "配送區域"] = np.nan
    n_region_filled = int(combo["配送區域"].notna().sum())
    log(f"  配送區域欄位（縣市＋區域）：{n_region_filled}／{len(combo)} 筆已可直接用於區域併單分組")

    # ---- 逐筆變更紀錄 ----
    changed_mask = original_addr.fillna("\x00__NA__") != combo[ADDRESS_COLUMN].fillna("\x00__NA__")
    for idx in combo.index[changed_mask]:
        change_log.append((
            combo.at[idx, "_原始列號"], ADDRESS_COLUMN, combo.at[idx, "配送地址_補值來源"],
            original_addr.at[idx], combo.at[idx, ADDRESS_COLUMN],
        ))
    log(f"  逐筆變更紀錄新增 {int(changed_mask.sum())} 筆")

    # ---- 清除暫存欄位，依 _原始列號 拆回原本兩個 DataFrame ----
    drop_cols = ["_addr_norm", "_ship_date_tmp", "_city", "_district", "_rest",
                 "_is_blank", "_is_full", "_is_placeholder", "_is_trusted", "_addr_code_key"]
    combo = combo.drop(columns=drop_cols)
    new_clean_df = (
        combo[combo["_原始列號"].isin(clean_keys)]
        .drop(columns=["_隔離原因"], errors="ignore")
        .reset_index(drop=True)
    )
    new_iso_df = combo[combo["_原始列號"].isin(iso_keys)].reset_index(drop=True)
    log("Part 18：配送地址補齊完成")

    return new_clean_df, new_iso_df, change_log


# ============================================================
# 3b. 離線「全國路名資料」回推／驗證（Part 25）
# ============================================================

def resolve_addresses_offline(clean_df: pd.DataFrame, iso_df: pd.DataFrame, change_log):
    """在規則補值（Part 18~23）之後，用內政部「全國路名資料」＋「村里清單」離線名冊
    補齊/驗證配送地址。完全不連外、100% 決定性、零個資外洩。名冊載入/唯一性索引/
    路名與村里擷取/交集回推見 services/address_reference.py。

    與規則補值相同，在「乾淨＋隔離」全體上合併計算（配送地址是門市屬性），清洗完依
    _原始列號 拆回兩個 DataFrame。去重後對唯一地址查一次、結果套回同地址所有列。

    加值而非覆蓋原則（受限於名冊為「非窮舉」清單，只做高精確度的事）：
      - 規則判『待確認』（配送區域留空）→ 依序嘗試「路名∩村里交集唯一 / 路名全國唯一 /
        村里全國唯一」回推；成功即補上配送區域、解除待確認、來源標記「全國路名資料回推」
        （離線比對='路名回推補齊'）。交集能救回路名與村里各自不唯一、但組合起來唯一的地址。
      - 規則已判出，且該地址的縣市+區域+路名確實存在於名冊 → 離線比對='路名相符(已驗證)'。
      - 其餘（路名/村里皆不唯一或不在名冊）→ 離線比對='未比對'，**不**視為錯誤、不改動
        規則結果（名冊非窮舉，查無不代表地址錯，避免大量假警報）。
    """
    from . import address_reference

    log("Part 25：離線全國路名資料 回推／驗證開始")
    clean_keys = set(clean_df["_原始列號"])
    iso_keys = set(iso_df["_原始列號"])
    combo = pd.concat([clean_df, iso_df], ignore_index=True, sort=False)

    # 預設欄位值（schema 固定，下游不必判斷欄位是否存在）。配送區域／補值來源保險轉 object，
    # 避免「整欄剛好全為缺值→float64」時用 .at 寫入字串在 pandas 3.0 下丟 dtype 錯誤。
    combo["配送區域"] = combo["配送區域"].astype(object)
    combo["配送地址_補值來源"] = combo["配送地址_補值來源"].astype(object)

    if not address_reference.is_loaded():
        combo["配送地址_離線比對"] = "未載入參考資料"
        log("Part 25：未載入全國路名資料（見 address_reference.py 取得說明），配送地址維持規則結果")
        new_clean_df = combo[combo["_原始列號"].isin(clean_keys)].reset_index(drop=True)
        new_iso_df = combo[combo["_原始列號"].isin(iso_keys)].reset_index(drop=True)
        return new_clean_df, new_iso_df, change_log

    combo["配送地址_離線比對"] = "未比對"

    # 去重：對唯一「最終配送地址字串」各算一次，再套回同字串所有列（地址是門市屬性，
    # 去重後量遠小於百萬列）。回傳 dict：addr -> (action, region_or_None)。
    query_addrs = combo[ADDRESS_COLUMN].dropna().astype(str)
    query_addrs = query_addrs[query_addrs.str.strip() != ""].unique().tolist()

    resolved = {}   # addr -> (city+district) 由唯一路名回推補齊者
    verified = set()  # addr -> 縣市+區域+路名皆存在於名冊者
    for addr in query_addrs:
        city, district, rest = split_city_district_rest(addr)
        if city and district:
            # 已有縣市+區域：只做正向驗證（存在於名冊才標已驗證，查無不視為錯誤）。
            # 傳入 rest（已去縣市/區域的路名+門牌），避免路名擷取誤含前綴。
            if address_reference.verify(city, district, rest):
                verified.add(addr)
        else:
            # 缺縣市或區域：用唯一路名回推（同樣傳入去除已知前綴後的 rest）
            region = address_reference.resolve_region(rest)  # 回傳 "縣市區域" 或 None（已 台/臺 正規化）
            if region:
                resolved[addr] = region

    # 效能：原本這裡對 combo 全體（可達百萬列）用 Python for 迴圈逐列 .at[] 存取，在正式資料集
    # 上實測是清洗流程最慢的一段（.at 逐列存取的常數開銷乘上百萬列）。resolved／verified 只是
    # 「地址字串 → 結果」的查表，改用向量化寫法（.map(dict) 與 .isin(set) 皆為 C 層雜湊查表）
    # 一次算出每列結果，再用布林遮罩批次寫回，結果與逐列版本完全等價，但快非常多。
    addr_str = combo[ADDRESS_COLUMN].fillna("").astype(str)
    region_col = combo["配送區域"]
    region_empty = region_col.isna() | (region_col == "")

    resolved_region = addr_str.map(resolved)              # 對應 resolved 的地址 → 回推區域字串，其餘 NaN
    fill_mask = resolved_region.notna() & region_empty     # 只在規則沒填出配送區域時才補
    verify_mask = (~resolved_region.notna()) & addr_str.isin(verified)

    n_fill = int(fill_mask.sum())
    n_verify = int(verify_mask.sum())

    if n_fill:
        combo.loc[fill_mask, "配送區域"] = resolved_region[fill_mask]
        combo.loc[fill_mask, "配送地址_待確認_flag"] = 0
        src_vals = combo.loc[fill_mask, "配送地址_補值來源"]
        combo.loc[fill_mask, "配送地址_補值來源"] = src_vals.apply(
            lambda s: (s + "；" if isinstance(s, str) and s else "") + "全國路名資料回推"
        )
        combo.loc[fill_mask, "配送地址_離線比對"] = "路名回推補齊"
    if n_verify:
        combo.loc[verify_mask, "配送地址_離線比對"] = "路名相符(已驗證)"

    log(f"  離線比對結果：唯一路名回推補齊 {n_fill} 筆／路名相符已驗證 {n_verify} 筆")

    new_clean_df = combo[combo["_原始列號"].isin(clean_keys)].reset_index(drop=True)
    new_iso_df = combo[combo["_原始列號"].isin(iso_keys)].reset_index(drop=True)
    log("Part 25：離線全國路名資料 回推／驗證完成")
    return new_clean_df, new_iso_df, change_log


# ============================================================
# 4. 新增「出貨日期」欄位（清洗文件.md Part 7）
# ============================================================

def add_shipdate_column(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    shipdate = df["系統出貨單號"].map(parse_shipno_date)
    cols = list(df.columns)
    ship_pos = cols.index("系統出貨單號")
    df.insert(ship_pos, "出貨日期", shipdate)
    return df


# ============================================================
# 5. 輸出
# ============================================================

def _move_column_after(df: pd.DataFrame, col: str, after_col: str) -> pd.DataFrame:
    """把 col 這一欄搬到緊接在 after_col 右邊，其餘欄位順序不變（Part 24）。"""
    cols = [c for c in df.columns if c != col]
    idx = cols.index(after_col)
    cols.insert(idx + 1, col)
    return df[cols]


def write_output(clean_df: pd.DataFrame, iso_df: pd.DataFrame, sku_dim_df: pd.DataFrame,
                  change_log, output_path: str):
    log(f"寫出結果檔案：{output_path}")

    clean_out = clean_df.drop(columns=["_原始列號", "訂購數量_數值"], errors="ignore")
    clean_out = add_shipdate_column(clean_out)

    iso_out = iso_df.rename(columns={"_原始列號": "Excel原始列號", "_隔離原因": "隔離主要原因"})
    iso_out = iso_out.drop(columns=["訂購數量_數值"], errors="ignore")
    iso_out["命中所有規則"] = iso_out["隔離主要原因"]
    iso_out = add_shipdate_column(iso_out)
    # 欄位順序：Excel原始列號, 出貨日期, 系統出貨單號, ...其餘14欄(扣除系統出貨單號)...,
    #          材積清洗結果欄位, 配送地址補值新增欄位(Part 18), 隔離主要原因, 命中所有規則
    front = ["Excel原始列號", "出貨日期", "系統出貨單號"]
    middle = [c for c in ORDER_COLUMNS if c != "系統出貨單號"] + MATERIAL_OUTPUT_COLUMNS + ADDRESS_EXTRA_COLUMNS
    tail = ["隔離主要原因", "命中所有規則"]
    iso_out = iso_out[front + middle + tail]

    # Part 24：「配送區域」欄改移到緊接在「配送地址」右邊，方便併單時直接參照，
    # 不必再往後找到材積/儲位欄位群組後面才看得到。
    clean_out = _move_column_after(clean_out, "配送區域", ADDRESS_COLUMN)
    iso_out = _move_column_after(iso_out, "配送區域", ADDRESS_COLUMN)

    lut_df = pd.DataFrame(CATEGORY_TABLE, columns=["商品類別ID", "商品類別代碼", "商品類別名稱", "來源"])

    chg_df = pd.DataFrame(change_log, columns=["Excel列號", "欄位", "規則", "原始值", "清理後值"])

    # 改用 openpyxl 的 write_only 串流模式手動逐列寫入（官方文件推薦的大檔案低記憶體寫法）。
    # 加上材積清洗欄位後，「乾淨資料」逼近104萬列×21欄（Part 11 起已將24欄逐維度flag/
    # 校正來源濃縮為1欄「材積待複核」，欄數比 Part 9/10 剛完成時的42欄大幅縮減；Part 18
    # 另外再加上「配送地址_補值來源」／「配送地址_待確認_flag」2欄，但仍維持串流寫入，
    # 理由如下，不因欄數變少而改回一般模式）：
    #   - 原本的 pd.ExcelWriter(engine="openpyxl") 一般模式，在本機（實測 7.7GB RAM）
    #     曾實際跑到寫檔案這一步直接記憶體存取違規當掉（exit code -1073741819）。
    #   - 改用 pd.ExcelWriter(engine="xlsxwriter", constant_memory=True) 雖不再當機，
    #     但實測發現嚴重的靜默資料遺失：寫出的儲存格幾乎全部變空白，只剩最左邊一欄有值，
    #     其餘41欄讀回來全是NaN——不會報錯，但資料是壞的，已放棄此做法。
    #   - write_only 模式：不建立完整的儲存格物件模型，逐列append、逐列flush到檔案，
    #     已用小範例（含缺值/日期/布林混合型別）驗證讀回結果與寫入前完全一致。
    from openpyxl import Workbook

    def _write_df_streaming(wb, sheet_name, df):
        ws = wb.create_sheet(sheet_name)
        ws.append(list(df.columns))
        clean_for_write = df.astype(object).where(df.notna(), None)
        for row in clean_for_write.itertuples(index=False, name=None):
            ws.append(row)

    wb = Workbook(write_only=True)
    _write_df_streaming(wb, "乾淨資料", clean_out)
    _write_df_streaming(wb, "隔離資料", iso_out)
    _write_df_streaming(wb, "商品類別對應表", lut_df)
    _write_df_streaming(wb, "逐筆變更紀錄", chg_df)
    _write_df_streaming(wb, "商品材積主檔", sku_dim_df)
    wb.save(output_path)

    return len(clean_out), len(iso_out)


# ============================================================
# main
# ============================================================


# ============================================================
# 後端服務入口（取代原本的 main()／CLI 執行方式）
# ------------------------------------------------------------
# 原腳本的 main() 是「讀檔案路徑→跑清洗→寫出 xlsx」的一次性 CLI 流程。
# 後端服務改成「接收上傳的 DataFrame→跑清洗→回傳 DataFrame／JSON」，
# 清洗邏輯本身（apply_isolation_rules／resolve_categories_and_reclassify／
# isolate_nondescriptive_names／clean_dimensions_and_weight／
# backfill_delivery_address／add_shipdate_column）完全不變，只是把
# 「輸入來源」與「輸出去處」從檔案系統換成記憶體物件，供 FastAPI 路由呼叫。
# ============================================================

from dataclasses import dataclass, field


def load_and_normalize_df(raw_df: pd.DataFrame):
    """load_and_normalize() 的記憶體版本：輸入已讀好的 DataFrame（而非檔案路徑），
    其餘正規化步驟（NULL 字樣→真空值、去空白、型別轉換、日期解析）完全相同。"""
    df = raw_df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    missing = [c for c in ORDER_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"上傳資料缺少必要欄位：{missing}；需包含 {ORDER_COLUMNS}")
    df = df[ORDER_COLUMNS].copy()
    df.insert(0, "_原始列號", range(2, len(df) + 2))
    original_row_count = len(df)
    log(f"原始資料列數（不含表頭）：{original_row_count}")

    for col in NULL_NORMALIZE_COLUMNS:
        df[col] = normalize_null_series(df[col])
    for col in TRIM_COLUMNS:
        df[col] = trim_series(df[col])
    df["商品ID"] = df["商品ID"].map(to_int_display)
    df["商品類別"] = df["商品類別"].map(to_int_display)
    df["訂購數量_數值"] = pd.to_numeric(df["訂購數量"], errors="coerce")
    return df, original_row_count


@dataclass
class CleaningResult:
    clean_df: pd.DataFrame
    iso_df: pd.DataFrame
    sku_dim_df: pd.DataFrame
    change_log_df: pd.DataFrame
    category_table_df: pd.DataFrame
    stats: dict = field(default_factory=dict)


def run_cleaning_pipeline(raw_df: pd.DataFrame) -> CleaningResult:
    """對應原 main() 的完整清洗流程（步驟順序完全比照原腳本，僅將檔案 I/O
    換成記憶體 DataFrame 傳遞）：
      1. 讀取與基礎正規化
      2. 八條優先序隔離規則，分流「乾淨資料」／「隔離資料」
      3. 商品類別關鍵字重分類（隔離資料救回乾淨資料）
      4. 商品名稱非描述性隔離
      5. 材積清洗（商品-長/寬/高/重量）＋建立商品材積主檔
      6. 配送地址補齊（縣市／區域）
      6b. 離線全國路名資料 回推／驗證（未載入名冊時自動略過）
      7. 加上「出貨日期」欄
    """
    log("========== 資料清洗開始（後端服務呼叫） ==========")
    df, original_row_count = load_and_normalize_df(raw_df)

    # 供規則2（完全重複列）比對用的「原始值」快照（正規化之前）
    raw_snapshot = raw_df.copy()
    raw_snapshot.columns = [str(c).strip() for c in raw_snapshot.columns]
    raw_snapshot = raw_snapshot[ORDER_COLUMNS].copy()

    df, change_log = apply_isolation_rules(df, raw_snapshot)

    clean_df = df.loc[df["_隔離原因"].isna()].drop(columns=["_隔離原因"]).copy()
    iso_df = df.loc[df["_隔離原因"].notna()].copy()
    log(f"初步分流：乾淨資料 {len(clean_df)} 筆／隔離資料 {len(iso_df)} 筆")

    clean_df, iso_df, reclass_stats = resolve_categories_and_reclassify(clean_df, iso_df)
    clean_df, iso_df, nondesc_moved = isolate_nondescriptive_names(clean_df, iso_df)

    total_after = len(clean_df) + len(iso_df)
    row_check_ok = (total_after == original_row_count)
    log(f"列數勾稽：{len(clean_df)} + {len(iso_df)} = {total_after}（原始 {original_row_count}）"
        + (" ✔ 一致" if row_check_ok else " ✘ 不一致，請檢查！"))

    clean_df, iso_df, sku_dim_df, change_log = clean_dimensions_and_weight(clean_df, iso_df, change_log)
    clean_df, iso_df, change_log = backfill_delivery_address(clean_df, iso_df, change_log)
    clean_df, iso_df, change_log = resolve_addresses_offline(clean_df, iso_df, change_log)

    # 比照 write_output()：乾淨資料丟掉內部輔助欄，加上「出貨日期」；隔離資料改名並補上
    # 「命中所有規則」欄（本移植版本沿用「隔離主要原因」做為命中規則展示，未逐條記錄多重命中）
    clean_out = clean_df.drop(columns=["_原始列號", "訂購數量_數值"], errors="ignore")
    clean_out = add_shipdate_column(clean_out)
    clean_out = _move_column_after(clean_out, "配送區域", ADDRESS_COLUMN)

    iso_out = iso_df.rename(columns={"_原始列號": "Excel原始列號", "_隔離原因": "隔離主要原因"})
    iso_out = iso_out.drop(columns=["訂購數量_數值"], errors="ignore")
    iso_out["命中所有規則"] = iso_out["隔離主要原因"]
    iso_out = add_shipdate_column(iso_out)
    iso_out = _move_column_after(iso_out, "配送區域", ADDRESS_COLUMN)

    change_log_df = pd.DataFrame(change_log, columns=["Excel列號", "欄位", "規則", "原始值", "清理後值"])
    category_table_df = pd.DataFrame(CATEGORY_TABLE, columns=["商品類別ID", "商品類別代碼", "商品類別名稱", "來源"])

    log(f"========== 清洗完成：乾淨資料 {len(clean_out)} 筆／隔離資料 {len(iso_out)} 筆 ==========")

    # 材積待複核：在「乾淨＋隔離」全體上計算（材積是商品物理屬性，跟該列是否被隔離無關），
    # 供 /api/clean/summary 的清洗成效摘要（準確度／待人工覆核）使用，對應前端原型
    # materialReviewCount（combo 全量加總，而非只算乾淨資料）。
    material_review_count = int(clean_out["材積待複核"].sum()) + int(iso_out["材積待複核"].sum())

    return CleaningResult(
        clean_df=clean_out,
        iso_df=iso_out,
        sku_dim_df=sku_dim_df,
        change_log_df=change_log_df,
        category_table_df=category_table_df,
        stats={
            "original_row_count": original_row_count,
            "clean_count": len(clean_out),
            "iso_count": len(iso_out),
            "row_check_ok": row_check_ok,
            "sku_count": len(sku_dim_df),
            "reclass": reclass_stats,
            "nondesc_moved": nondesc_moved,
            "material_review_count": material_review_count,
        },
    )


def export_cleaned_xlsx(result: "CleaningResult", output_path: str) -> None:
    """選用工具函式：把清洗結果寫成 xlsx（沿用原 write_output() 的分頁與串流寫入方式），
    供「下載清洗結果」端點使用。與原腳本差別只在於：DataFrame 已經是清洗完成的結果，
    不需要再呼叫 add_shipdate_column／_move_column_after（run_cleaning_pipeline 已處理過）。"""
    from openpyxl import Workbook

    def _write_df_streaming(wb, sheet_name, df):
        ws = wb.create_sheet(sheet_name)
        ws.append(list(df.columns))
        clean_for_write = df.astype(object).where(df.notna(), None)
        for row in clean_for_write.itertuples(index=False, name=None):
            ws.append(row)

    wb = Workbook(write_only=True)
    for sheet_name, df in _export_frames(result):
        _write_df_streaming(wb, sheet_name, df)
    wb.save(output_path)


EXPORT_SHEET_ORDER = ("乾淨資料", "隔離資料", "商品類別對應表", "逐筆變更紀錄", "商品材積主檔")


def _export_frames(result: "CleaningResult"):
    """匯出用的 (分頁名稱, DataFrame) 序列，xlsx 與 ZIP 兩種格式共用同一份定義與順序。"""
    by_name = {
        "乾淨資料": result.clean_df,
        "隔離資料": result.iso_df,
        "商品類別對應表": result.category_table_df,
        "逐筆變更紀錄": result.change_log_df,
        "商品材積主檔": result.sku_dim_df,
    }
    return [(name, by_name[name]) for name in EXPORT_SHEET_ORDER]


def export_cleaned_zip(result: "CleaningResult", output_path: str) -> None:
    """把清洗結果寫成 ZIP，內含 5 個 CSV（檔名對應原本 xlsx 的 5 個工作表）。

    為什麼不用 xlsx：openpyxl 是逐格建立儲存格再逐列 append，量大時非常慢。同一份正式資料
    （104 萬列＋隔離 7,808＋變更紀錄 10.3 萬＋材積主檔 5,699＋類別表 69）實測：
        xlsx  481 秒 / 121.7 MB
        ZIP    25 秒 /  28.3 MB      快約 19 倍、檔案小約 4 倍
    代價是 CSV 不帶型別資訊（見下方「說明.txt」對前導 0 的提醒）。

    寫法上刻意用 zf.open(...) ＋ TextIOWrapper 串流寫入，而不是先 df.to_csv() 取得整個
    字串再塞進 ZIP——後者會把百萬列的 CSV 全文一次放進記憶體（可達數百 MB），在資料量大時
    是不必要的風險。編碼用 utf-8-sig（BOM 由 TextIOWrapper 自動寫入），Excel 直接雙擊
    開啟中文不會變亂碼；換行固定 CRLF，與系統其他 CSV 下載一致。
    """
    import io
    import zipfile

    frames = _export_frames(result)
    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for name, df in frames:
            with zf.open(f"{name}.csv", "w") as raw:
                with io.TextIOWrapper(raw, encoding="utf-8-sig", newline="") as fh:
                    df.to_csv(fh, index=False, lineterminator="\r\n")
        # 說明檔用 CRLF，Windows 上用最陽春的記事本開也會正常分行。
        zf.writestr("說明.txt", _export_readme(frames).replace("\n", "\r\n").encode("utf-8-sig"))


def _export_readme(frames) -> str:
    """ZIP 內的說明檔：交代每個 CSV 對應哪個分頁、各有幾列，以及用 Excel 開啟的注意事項。"""
    lines = [
        "清洗結果匯出（CSV 版）",
        "",
        "本壓縮檔內的 5 個 CSV，對應先前 Excel 版本的 5 個工作表：",
    ]
    for name, df in frames:
        lines.append(f"    {name}.csv    {len(df):,} 列 × {len(df.columns)} 欄")
    lines += [
        "",
        "編碼：UTF-8（含 BOM），換行 CRLF。用 Excel 直接雙擊開啟，中文不會變亂碼。",
        "",
        "【用 Excel 開啟前請注意】",
        "CSV 不像 xlsx 會保存欄位型別，Excel 雙擊開啟時會自行猜測型別，可能造成：",
        "  ・以 0 開頭的代碼（例如部分商品編號、門市代碼）前導 0 被吃掉",
        "  ・長數字被轉成科學記號",
        "  ・看起來像日期的文字被轉成日期",
        "若要保留原樣，請改用「資料 → 取得資料 → 從文字/CSV」匯入，並把這些欄位指定為「文字」。",
    ]
    return "\n".join(lines)


# ============================================================
# 更新記錄（簡化版，依時間序；完整實測數據與逐項規則細節見各 Part
# 對應常數/函式旁的內部註解，例如 VOLUME_EXTREME_MAX、
# SHARED_ADDR_SPARSE_MAX_ROWS_PER_CODE 等定義處）
# ============================================================
# Part 1~9　：依清洗文件.md 建立基礎清洗流程（NULL正規化、8條隔離規則、商品類別
#            補正、材積清洗等），已用原始檔逐列比對驗證與人工清洗結果相符。
# Part 10　：材積清洗新增「長寬高相乘體積」極端值門檻(VOLUME_EXTREME_MAX)、
#            共用可疑值偵測(SHARED_VALUE_MIN_DISTINCT_SKU)、測試資料商品名稱
#            關鍵字排除(TEST_DATA_NAME_PATTERNS)。
# Part 11　：材積清洗輸出欄位精簡為單一「材積待複核」旗標；隔離資料關鍵字重分類
#            前補檢查數量/單位規則，避免誤救回乾淨資料（實測抓到7筆）。
# Part 12　：主控台輸出改為UTF-8，避免Windows非UTF-8語系下log符號造成腳本崩潰。
# Part 13　：類別校正候選值須同時通過該欄極端值門檻才採用，避免超標值被誤標為
#            已校正、靜默混入乾淨資料。
# Part 14　：Excel讀取改用 read_excel_robust()，優先 calamine、失敗退回 openpyxl，
#            解決700MB大檔案 openpyxl 解析失敗問題（需 pip install python-calamine）。
# Part 15　：長寬高缺值改用「同商品類別中位數」推估；新增「材積來源」欄；
#            材積待複核改為只看長寬高，不含重量。
# Part 16　：新增「儲位分類」（貨架/棧板），依材積>50L或最長邊>60cm判定，
#            供下游 ABC分組與儲位配置使用。
# Part 17　：非描述性商品名稱（純規格代號，如「(60) 10/15」）隔離待人工補正式品名。
# Part 18　：配送地址補齊縣市/區域——舊制行政區正規化、共用預設值偵測、同門市代碼
#            可信版本回填；新增「配送地址_補值來源」／「配送地址_待確認_flag」。
#            實測（2026-08-06）：查無可信版本352,968筆待業務端補門市地址主檔，
#            其餘皆已回填（新舊制更名11,503筆／單一版本回填43,312筆／依日期多版本
#            回填3,447筆／共用預設值排除14筆）。
# Part 19　：新增依賴套件自舉(_bootstrap_dependencies())。原因：本機 Python 由
#            uv 管理，屬 externally-managed 環境，直接 pip install 會被擋下，
#            每次「從頭」在新環境執行都要手動排查。改為腳本啟動時自動偵測缺套件，
#            自動建立 .venv 並安裝好 pandas/openpyxl/python-calamine 後自我重啟，
#            不需再手動下 pip 指令。重啟改用 subprocess.run() 而非 os.execv()：
#            實測 os.execv() 在本機含全形/空白字元路徑下會把路徑重複拼接、找不到
#            檔案，subprocess.run() 無此問題。
# Part 20　：新增「執行清洗.bat」一鍵啟動器＋修正用法說明。原因：本機終端機打裸的
#            python 指令對應到 Windows 市集的假殼，非互動執行下完全沒有反應
#            （不報錯也不執行），使用者誤以為腳本沒有作用。雙擊 .bat 會自動改用
#            .venv（或首次執行時已知可用的 uv 直譯器）呼叫本腳本，並可拖曳/貼上
#            輸入檔路徑，不需要再記指令或處理 python 別名問題。
# Part 21　：新增「配送區域」欄（供區域併單策略直接分組使用）。因應使用者反映：原始
#            地址有些完全空白、有些只有路名門牌（無縣市/區域），若直接依完整地址字串
#            分組，同一區域會因寫法不同被拆成多組，無法正確併單。Part 18 的補值邏輯
#            已可處理這兩種情況（同門市代碼跨時間互相佐證回填），本次新增的「配送區域」
#            欄則是在補值完成後，只取縣市＋區域（例：新北市內湖區），不含路名門牌，
#            供下游併單邏輯直接依此欄分組；仍查無可信縣市/區域者（配送地址_待確認_flag
#            =1 的列）此欄留空，需人工確認門市地址後才能併單。
# Part 22　：新增地址前綴回推機制，找回一批先前被標記「查無可信版本、待確認」但其實
#            有跡可循的地址（新增 normalize_address_prefix()，依序處理，詳見函式與
#            TOWNSHIP_TO_COUNTY／CITY_ABBREVIATIONS 內部註解）：
#              1. 開頭誤帶郵遞區號（3~6碼數字）：去除後才進行縣市比對。
#              2. 常見城市簡寫（例：北市→台北市）：展開為正式縣市名稱。
#              3. 省略縣名、只寫全國唯一（無同名歧義）鄉/鎮/縣轄市名稱者（例：屏東市、
#                 琉球鄉這類離島/縣轄市地名）：依 TOWNSHIP_TO_COUNTY 回推補上所屬縣。
#                 直轄市/省轄市轄下的「區」不在此列——區名在不同城市間常重複（如中山
#                 區同時存在於台北市、高雄市等），無法安全回推，仍維持人工確認。
#            這批地址回推後即可直接判定為完整可信地址，即使該門市代碼只有這一筆、沒有
#            其他列可互相佐證，也不需要再等 Part 18 的同代碼跨列佐證邏輯。已用單元測試
#            覆蓋四種案例（屏東市、琉球鄉、郵遞區號前綴、城市簡寫）及既有案例回歸測試，
#            全數通過；過程中並抓到一個真實 bug：「新北」同時是「新北市」的簡寫也是它
#            的字面前綴，若不先排除已完整的地址，「新北市...」會被誤展開成「新北市
#            市...」，已在 expand_city_abbreviation() 加上「已是完整縣市名稱則略過」
#            的檢查修正。複核真實資料時再抓到兩個問題並修正：
#              a. 既有 Part 18 的 DISTRICT_RE 對「前鎮區」（高雄）、「平鎮區」（桃園）、
#                 「左鎮區」（台南）這三個區名會誤判——因為名稱中間剛好也有「鎮」字，
#                 非貪婪比對提前於「前鎮」處收尾，漏掉「區」字（實測受影響 37,429+15
#                 筆）。新增 SPECIAL_DISTRICT_NAMES 特例清單，split_city_district_rest()
#                 比對前先攔截這幾個已知特例。
#              b. 「配送區域」欄先前只依「地址文字表面能否解析出縣市+區域」決定是否
#                 填值，沒有同步檢查 配送地址_待確認_flag，導致命中共用預設值偵測（例：
#                 「高雄市橋頭區新莊村大學南路1號」，文字完整但系統判斷是跨門市代碼
#                 共用的可疑預設值）的列，明明待確認=1，配送區域卻仍顯示看似可用的
#                 區域，可能被下游併單邏輯誤用。已改為 待確認_flag=1 者「配送區域」
#                 強制留空，兩者保持一致。
# Part 23　：因應使用者提供的三個實際案例，再擴充兩種地址前綴回推規則（新增
#            OLD_TOWNSHIP_TO_NEW_DISTRICT／OLD_COUNTY_ABBREVIATIONS，詳見內部註解）：
#              1. 省略舊制縣名、直接寫「已併入直轄市的舊制鄉鎮市名稱」者（例：「鳳山市
#                 海洋一路121號」省略了「高雄縣」）：新增 infer_new_district_from_old_
#                 township()，依 OLD_TOWNSHIP_TO_NEW_DISTRICT（涵蓋舊高雄/台南/台北/
#                 台中/桃園縣共121個已改制鄉鎮市）回推轉換為現制「直轄市＋區」寫法
#                 （鳳山市→高雄市鳳山區）。舊桃園縣治「桃園市」因與現制桃園市同名，
#                 刻意不收錄，避免誤判。
#              2. 舊制縣名簡寫（例：「高縣茄萣鄉...」的「高縣」）：新增
#                 expand_old_county_abbreviation()／OLD_COUNTY_ABBREVIATIONS，展開為
#                 完整舊制縣名後交給既有 normalize_old_county_address() 接手轉換。
#            實作時把 TOWNSHIP_TO_COUNTY／OLD_TOWNSHIP_TO_NEW_DISTRICT 的比對方式，從
#            Part 22 原本「非貪婪正規表示式配到市/鎮/鄉即收尾、再查表」，改為「已知名稱
#            清單依長度由長至短建成一個比對用正規表示式，一次比對」——原因是複核時發現
#            這個新表裡有「新市鄉」「左鎮鄉」「平鎮市」三個名稱，中間剛好也有市/鎮字元，
#            用舊比對方式一樣會被提前收尾誤判（跟 Part 22 修正 DISTRICT_RE 的「前鎮區」
#            是同一類問題）；改成已知清單整體比對後，這類問題不會再發生，且比對的清單
#            本身是封閉、固定的，效能沒有實質影響。已用單元測試覆蓋這三個易錯名稱，
#            以及兩個使用者提供的實際案例，全數通過。
#            使用者另提出第三種案例（「光榮里光榮街90號1樓」依路名推估為「澎湖縣馬公市
#            光榮里光榮街90號1樓」），評估後**未**採用同樣方式做成通用規則：TOWNSHIP_TO_
#            COUNTY／OLD_TOWNSHIP_TO_NEW_DISTRICT 這兩類回推之所以安全，是因為鄉/鎮/
#            縣轄市層級的行政區名稱在全國範圍內具有官方保證的唯一性（不會有兩個縣有
#            同名鄉鎮市），但「街道名稱」沒有這種保證——「光榮街」這類街名在台灣各地
#            重複出現非常普遍，若做成通用規則會有很高機率把其他縣市裡真實存在的同名
#            街道地址誤改成澎湖縣馬公市，屬於資料正確性風險，故此案例未展開為通用機制，
#            留待使用者確認或改以人工個案處理。（後續實際上網查證：「光榮里」至少
#            同時存在於澎湖縣馬公市與高雄市鼓山區兩地，「光榮街/路」也同時存在於苗栗縣
#            竹南鎮、宜蘭縣羅東鎮，證實兩者在全國都不是唯一，此案例確定不適合做成通用
#            規則，維持原判斷。）
# Part 24　：「配送區域」欄位置從欄位群組最後段，改移到緊接在「配送地址」右邊
#            （新增 _move_column_after() 輔助函式，寫檔前套用在乾淨資料／隔離資料），
#            方便併單時在同一視野直接對照地址與區域，不必再往後翻到材積/儲位欄位群組
#            之後才看得到。
# Part 25　：用內政部「全國路名資料」＋「村里清單」離線名冊補齊/驗證配送地址（新增
#            services/address_reference.py＋resolve_addresses_offline()）。因應使用者
#            需求：想借助地名資訊讓配送地址更完整，且明確選擇「離線」方案——完全不連外、
#            100% 決定性、零個資外洩（相對於線上地理編碼服務會有個資外送、結果隨時間
#            變動、每日配額等問題）。
#            資料來源（皆政府資料開放授權條款第1版，可自由使用/散布，隨專案打包快照）：
#              - 路名：戶政司「全國路名資料」ODRP049，欄位 city／site_id(縣市+區域)／road。
#                113年版 35,223 筆、21,447 個不同路名，17,277 個(80.6%)全國唯一。
#              - 村里：國土測繪中心 NLSC ListCounty/ListTown/ListVillage 彙整，欄位
#                city／district／village。7,667 個村里、5,090 個不同名、4,006 個(78.7%)全國唯一。
#            可用環境變數 WAREHOUSE_ROAD_DATA／WAREHOUSE_VILLAGE_DATA 指向更新版。
#            設計要點：
#              1. 決定性/隱私：純本地字典查詢，無網路、無外部相依、無隨機性，天然符合
#                 本管線「相同輸入結果完全一致」原則，也不外送任何客戶地址。
#              2. 高精確度回推（名冊為非窮舉清單，不宜反推「查無=錯」）。規則判『待確認』時
#                 依序：(A) 路名∩村里交集唯一 →(B) 路名全國唯一 →(C) 村里全國唯一，成功即
#                 回推補上縣市+區域、解除待確認、來源標記「全國路名資料回推」。(A) 交集能救回
#                 路名與村里各自不唯一（中山路遍地都是、某某里也重名）但組合起來唯一的一大批
#                 地址，是加入村里資料的主要效益。已判出且縣市+區域+路名存在於名冊→標
#                 『路名相符(已驗證)』。其餘標『未比對』，不改動規則結果。新增欄位：配送地址_離線比對。
#              3. 先天限制（誠實揭露）：路名/村里都不唯一、交集也無法收斂到單一時仍待確認；
#                 空白/非地址文字無從判斷。命中率低於線上模糊比對，換得零外洩＋完全決定性，
#                 此為使用者明確選擇的取捨。路名擷取取結尾為 路/街/道/大道 的詞去段別、村里
#                 取開頭結尾為 村/里 的詞；名冊「臺」字正規化為管線慣用「台」字，避免補齊值
#                 與其他配送區域 key 不一致害併單分組被拆開。
#              4. 不阻塞、可退場：兩份名冊都缺時 is_loaded()=False，此步驟略過、配送地址
#                 維持規則結果，/api/clean/run 照常完成，對現有部署零破壞。
#            實作只用 Python 標準函式庫（csv/gzip），未新增相依套件。已用單元測試（交集/
#            路名唯一/村里唯一三路徑、常見路名不回推、已判出者驗證、未載入名冊 no-op）＋
#            全管線 smoke test 驗證，全數通過。
#            （TGOS 線上/批次串接曾在此位置實作過，後依使用者選擇改為純離線方案而移除；
#            若日後要再提高覆蓋率，可再加完整門牌點位資料。）
